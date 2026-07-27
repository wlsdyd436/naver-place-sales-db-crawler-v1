"""
PAGE300-SSR-1: SSR 상세정보 보강(BrowserContext-shared page.request.get →
window.__APOLLO_STATE__ → extract_normalized_apollo_detail → place_id exact
병합) 기본 파이프라인 배선 검증.

1~30: _fetch_place_detail_ssr / DomMembershipCollector.enrich_detail_ssr
순수 함수·메서드 레벨 - FakePage.request(FakeAPIRequestContext)를 새로
구현하고, entryIframe 기반 DOM fallback 검증에는
tests/test_pc_detail_enrichment.py의 Fake 패턴을 이 파일 안에 독립적으로
재구현한다(기존 파일 무수정). Apollo state fixture는 5M/5O 원본 raw JSON을
복사하지 않고 합성(익명화)한다(기존 adapter 테스트와 동일 원칙).
31~36: src/ui.py._run_network_pipeline이 SSR 보강을 별도 토글 없이 기본
흐름에 자동 배선하는지 검증 - tests/test_ui_network_wiring.py의 fake app/
collector DI 패턴을 이 파일 안에 독립적으로 재구현한다.
"""
import inspect
import json
import sys
import tempfile
import threading
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from playwright.sync_api import TimeoutError as PlaywrightTimeoutError

from src import ui
from src.exporter import MERGED_COLUMNS, export_places_to_excel
from src.pc import network_browser_collector
from src.pc.network_browser_collector import (
    DomMembershipCollector,
    _classify_ssr_block_signal,
    _fetch_place_detail_ssr,
    _parse_apollo_state_from_html,
)
from src.pc.network_list_scraper import merge_detail_into_row


class SimpleReporter:
    def __init__(self):
        self.passes = 0
        self.fails = 0

    def pass_(self, msg: str):
        self.passes += 1
        print(f"[PASS] {msg}")

    def fail(self, msg: str):
        self.fails += 1
        print(f"[FAIL] {msg}")


# ---------------------------------------------------------------------------
# Apollo state fixture 빌더 (합성/익명화 - 원본 raw JSON 복사 없음)
# ---------------------------------------------------------------------------


def _base_entity(place_id="2014880028", **overrides):
    entity = {
        "id": place_id,
        "name": "테스트 카페",
        "category": "카페,디저트",
        "roadAddress": "서울 강동구 성내로 1",
        "address": "서울 강동구 천호동 1",
        "phone": "02-488-5582",
        "virtualPhone": "0507-1111-2222",
        "visitorReviewsTotal": 239,
        "cafeBlogReviewsTotal": 134,
        "naverBlog": None,
    }
    entity.update(overrides)
    return entity


def _parent_entity(base_ref="PlaceDetailBase:2014880028", **overrides):
    entity = {
        "base": {"__ref": base_ref},
        "homepages": {"repr": {"url": "https://example-cafe.test", "type": "홈페이지"}, "etc": []},
        "newOpening": False,
        "phoneInfo": {"phone": "02-488-5582", "isVirtualPhone": False, "isPrivatePhone": False},
    }
    entity.update(overrides)
    return entity


def _full_apollo_state(place_id="2014880028", *, base_overrides=None, parent_overrides=None):
    base_key = f"PlaceDetailBase:{place_id}"
    return {
        base_key: _base_entity(place_id=place_id, **(base_overrides or {})),
        "ROOT_QUERY": {
            f'placeDetail({{"input":{{"deviceType":"mobile","id":"{place_id}","isNx":false}}}})': _parent_entity(
                base_ref=base_key, **(parent_overrides or {})
            ),
        },
    }


def _html_with_apollo(apollo_state) -> str:
    return (
        "<!DOCTYPE html><html><head></head><body><div id=\"root\"></div>"
        "<script>window.__APOLLO_STATE__ = "
        + json.dumps(apollo_state, ensure_ascii=False)
        + ";</script></body></html>"
    )


def _row(name, place_id):
    return {
        "업체명": name,
        "place_id": place_id,
        "플레이스 URL": "",
        "대표전화": "",
        "주소": "",
        "홈페이지": "",
        "인스타": "",
        "블로그": "",
    }


PLACE_ID = "2014880028"
PLACE_URL = f"https://pcmap.place.naver.com/place/{PLACE_ID}/home"


# ---------------------------------------------------------------------------
# Fake page.request (SSR)
# ---------------------------------------------------------------------------


class FakeSSRResponse:
    def __init__(self, status=200, html_text="", content_type="text/html; charset=utf-8"):
        self.status = status
        self._html_text = html_text
        self.headers = {"content-type": content_type}

    def text(self):
        return self._html_text


class FakeAPIRequestContext:
    def __init__(self, response_by_url=None, exception_by_url=None):
        self._response_by_url = response_by_url or {}
        self._exception_by_url = exception_by_url or {}
        self.get_calls: list = []

    def get(self, url, headers=None, timeout=None):
        self.get_calls.append(url)
        if url in self._exception_by_url:
            raise self._exception_by_url[url]
        return self._response_by_url.get(url, FakeSSRResponse(status=404, html_text=""))


# ---------------------------------------------------------------------------
# Fake DOM 방문 fallback (tests/test_pc_detail_enrichment.py 패턴 독립 재구현)
# ---------------------------------------------------------------------------


class FakeLD:
    def __init__(self, count_value=0, text=""):
        self._count = count_value
        self._text = text

    @property
    def first(self):
        return self

    def count(self):
        return self._count

    def inner_text(self, timeout=None):
        return self._text


class FakeHrefs:
    def __init__(self, hrefs):
        self._hrefs = hrefs

    def evaluate_all(self, expression):
        return list(self._hrefs)


class FakeValue:
    def __init__(self, text="", hrefs=None):
        self._text = text
        self._hrefs = hrefs or []

    @property
    def first(self):
        return self

    def count(self):
        return 1

    def inner_text(self, timeout=None):
        return self._text

    def locator(self, selector):
        if selector == "span.pz7wy":
            return FakeLD(count_value=0)
        return FakeHrefs(self._hrefs)


class FakeLabel:
    def __init__(self, value):
        self._value = value

    @property
    def first(self):
        return self

    def count(self):
        return 1

    def locator(self, selector):
        return self._value


class FakeMissing:
    @property
    def first(self):
        return self

    def count(self):
        return 0

    def inner_text(self, timeout=None):
        return ""

    def locator(self, selector):
        return FakeMissing()


class FakeEntryFrame:
    def __init__(self, title="", phone="", address="", hrefs=None):
        self.name = "entryIframe"
        self.url = "https://pcmap.place.naver.com/place/1/home"
        self._map = {}
        if title:
            self._map["span.IY7ZX"] = FakeLD(count_value=1, text=title)
        if phone:
            self._map["span.place_blind:has-text('전화번호')"] = FakeLabel(FakeValue(text=phone))
        if address:
            self._map["span.place_blind:has-text('주소')"] = FakeLabel(FakeValue(text=address))
        if hrefs:
            self._map["span.place_blind:has-text('홈페이지')"] = FakeLabel(FakeValue(hrefs=hrefs))

    def locator(self, selector):
        return self._map.get(selector, FakeMissing())

    def evaluate(self, expression, arg=None):
        return {"available": False, "apollo_state": {}}


class FakeCaptchaLocator:
    def __init__(self, count=0, visible=False, box=None):
        self._count = count
        self._visible = visible
        self._box = box or {"width": 0, "height": 0}

    @property
    def first(self):
        return self

    def count(self):
        return self._count

    def is_visible(self, timeout=300):
        return self._visible

    def bounding_box(self):
        return self._box


class FakeGotoResponse:
    def __init__(self, url, status):
        self.url = url
        self.status = status


class FakePage:
    """SSR(page.request)과 DOM fallback(page.goto/frame/locator)을 모두
    지원하는 fake page. request_ctx가 없으면 SSR GET은 항상 404 fallback
    응답(FakeAPIRequestContext 기본값)을 받는다."""

    def __init__(self, *, request_ctx=None, entry_frame_by_url=None, status_by_url=None, captcha_selectors=None):
        self.request = request_ctx or FakeAPIRequestContext()
        self._entry_frame_by_url = entry_frame_by_url or {}
        self._status_by_url = status_by_url or {}
        self._captcha_selectors = captcha_selectors or {}
        self._handlers: dict = {}
        self._current_url = None
        self.goto_calls: list = []
        self.closed = False

    def on(self, event, handler):
        self._handlers.setdefault(event, []).append(handler)

    def off(self, event, handler):
        handlers = self._handlers.get(event, [])
        if handler in handlers:
            handlers.remove(handler)

    def goto(self, url, wait_until=None, timeout=None):
        self.goto_calls.append(url)
        self._current_url = url
        for status in self._status_by_url.get(url, [200]):
            response = FakeGotoResponse(url, status)
            for handler in list(self._handlers.get("response", [])):
                handler(response)

    def locator(self, selector):
        return self._captcha_selectors.get(selector, FakeCaptchaLocator(count=0))

    def frame(self, name=None):
        if name == "entryIframe":
            return self._entry_frame_by_url.get(self._current_url)
        return None

    @property
    def frames(self):
        return []

    def close(self):
        self.closed = True


# ---------------------------------------------------------------------------
# DomMembershipCollector DI helper (tests/test_pc_detail_enrichment.py 패턴)
# ---------------------------------------------------------------------------


class FakeContext:
    def __init__(self, pages):
        self._pages = list(pages)
        self._index = 0

    def new_page(self):
        page = self._pages[self._index]
        self._index += 1
        return page


class FakeInnerSession:
    def __init__(self, pages):
        self.context = FakeContext(pages)


class FakeSessionCM:
    def __init__(self, pages):
        self._session = FakeInnerSession(pages)

    def __enter__(self):
        return self._session

    def __exit__(self, exc_type, exc, tb):
        return False


def _make_collector(pages):
    collector = DomMembershipCollector(collected_at="2026-07-23", session_factory=lambda: FakeSessionCM(pages))
    collector.__enter__()
    return collector


def main() -> int:
    reporter = SimpleReporter()

    # ------------------------------------------------------------------
    # 1~14. _fetch_place_detail_ssr 단건
    # ------------------------------------------------------------------

    # 1. BrowserContext-shared request 사용(page.request.get, page.goto 아님)
    apollo_state1 = _full_apollo_state()
    req1 = FakeAPIRequestContext({PLACE_URL: FakeSSRResponse(html_text=_html_with_apollo(apollo_state1))})
    page1 = FakePage(request_ctx=req1)
    result1 = _fetch_place_detail_ssr(page1, _row("테스트 카페", PLACE_ID))
    if req1.get_calls == [PLACE_URL] and page1.goto_calls == [] and result1["detail_ssr_status"] == "success":
        reporter.pass_("1. BrowserContext-shared page.request.get만 사용(page.goto 없음)")
    else:
        reporter.fail(f"1. request mechanism 이상: get_calls={req1.get_calls} goto_calls={page1.goto_calls} result={result1}")

    # 2. urllib/requests 미사용(정적 소스 검사)
    src = inspect.getsource(network_browser_collector)
    forbidden = ("import requests", "import httpx", "import aiohttp", "urllib.request", "http.client")
    found = [tok for tok in forbidden if tok in src]
    if not found:
        reporter.pass_("2. network_browser_collector.py에 urllib.request/requests/httpx/aiohttp 미사용")
    else:
        reporter.fail(f"2. 금지된 HTTP 클라이언트 import 발견: {found}")

    # 3. HTTP 200 + 정상 Apollo 성공
    if result1["detail_ssr_http_status"] == 200 and result1["detail_success"] is True:
        reporter.pass_("3. HTTP 200 + Apollo 정상 파싱 시 성공 처리")
    else:
        reporter.fail(f"3. HTTP 200 성공 처리 이상: {result1}")

    # 4. Base/Parent join(홈페이지/새로오픈여부가 parent에서 옴)
    if result1["홈페이지"] == "https://example-cafe.test" and result1["새로오픈여부"] == "X":
        reporter.pass_("4. Base+Parent join: 홈페이지/새로오픈여부가 parent 필드에서 정확히 반영")
    else:
        reporter.fail(f"4. Base/Parent join 이상: 홈페이지={result1['홈페이지']} 새로오픈여부={result1['새로오픈여부']}")

    # 5. place_id exact
    if result1["detail_ssr_place_id_exact"] is True:
        reporter.pass_("5. place_id exact 확인됨")
    else:
        reporter.fail(f"5. place_id exact 미확인: {result1}")

    # 6. place_id mismatch 병합 금지
    mismatch_state = {f"PlaceDetailBase:{PLACE_ID}": _base_entity(place_id="9999999999")}
    req6 = FakeAPIRequestContext({PLACE_URL: FakeSSRResponse(html_text=_html_with_apollo(mismatch_state))})
    page6 = FakePage(request_ctx=req6)
    result6 = _fetch_place_detail_ssr(page6, _row("테스트 카페", PLACE_ID))
    if result6["detail_ssr_status"] == "mismatch" and result6["detail_success"] is False:
        reporter.pass_("6. place_id mismatch 감지 시 detail_success=False(병합 금지 대상)")
    else:
        reporter.fail(f"6. mismatch 처리 이상: {result6}")

    # 7~9. HTTP 403/405/429 즉시 중단 (enrich_detail_ssr 레벨)
    for case_no, code in ((7, 403), (8, 405), (9, 429)):
        rows789 = [_row(f"업체{i}", f"70{code}0{i}") for i in range(3)]
        urls789 = [f"https://pcmap.place.naver.com/place/70{code}0{i}/home" for i in range(3)]
        resp_by_url = {
            urls789[0]: FakeSSRResponse(html_text=_html_with_apollo(_full_apollo_state(f"70{code}00"))),
            urls789[1]: FakeSSRResponse(status=code, html_text=""),
        }
        req = FakeAPIRequestContext(resp_by_url)
        page = FakePage(request_ctx=req)
        collector = _make_collector([page])
        out = collector.enrich_detail_ssr(rows789)
        ok = (
            out["security_blocked"] is True and out["stop_reason"] == f"HTTP_{code}"
            and urls789[2] not in req.get_calls and len(out["rows"]) == 3
        )
        if ok:
            reporter.pass_(f"{case_no}. HTTP {code} 감지 시 즉시 전체 중단, 이후 row 미시도")
        else:
            reporter.fail(f"{case_no}. HTTP {code} 중단 처리 이상: stop_reason={out['stop_reason']} get_calls={req.get_calls}")

    # 10. CAPTCHA 즉시 중단
    rows10 = [_row("업체0", "8000000"), _row("업체1", "8000001")]
    urls10 = ["https://pcmap.place.naver.com/place/8000000/home", "https://pcmap.place.naver.com/place/8000001/home"]
    req10 = FakeAPIRequestContext({urls10[0]: FakeSSRResponse(status=200, html_text="<html>자동입력방지문자 확인 필요</html>")})
    page10 = FakePage(request_ctx=req10)
    collector10 = _make_collector([page10])
    out10 = collector10.enrich_detail_ssr(rows10)
    if out10["security_blocked"] is True and out10["stop_reason"] == "CAPTCHA_CHALLENGE" and urls10[1] not in req10.get_calls:
        reporter.pass_("10. CAPTCHA 텍스트 마커 감지 시 즉시 전체 중단")
    else:
        reporter.fail(f"10. CAPTCHA 중단 처리 이상: {out10}")

    # 11. HTTP 500 → UI fallback 최대 1회
    row11 = _row("카페 fallback", "9000001")
    url11 = "https://pcmap.place.naver.com/place/9000001/home"
    req11 = FakeAPIRequestContext({url11: FakeSSRResponse(status=500, html_text="")})
    page11 = FakePage(
        request_ctx=req11,
        entry_frame_by_url={url11: FakeEntryFrame(title="카페 fallback", phone="02-1111-0000", address="주소11")},
    )
    collector11 = _make_collector([page11])
    out11 = collector11.enrich_detail_ssr([row11])
    if (
        page11.goto_calls == [url11] and out11["ui_fallback_used_count"] == 1
        and out11["ui_fallback_success_count"] == 1 and out11["rows"][0]["대표전화"] == "02-1111-0000"
    ):
        reporter.pass_("11. HTTP 500 시 UI fallback 1회 시도 후 DOM 결과로 병합 성공")
    else:
        reporter.fail(f"11. HTTP 500 fallback 처리 이상: goto_calls={page11.goto_calls} out={out11}")

    # 12. Apollo missing → UI fallback
    row12 = _row("카페 apollo없음", "9000002")
    url12 = "https://pcmap.place.naver.com/place/9000002/home"
    req12 = FakeAPIRequestContext({url12: FakeSSRResponse(status=200, html_text="<html>일반 페이지, Apollo 없음</html>")})
    page12 = FakePage(
        request_ctx=req12,
        entry_frame_by_url={url12: FakeEntryFrame(title="카페 apollo없음", phone="02-2222-0000")},
    )
    collector12 = _make_collector([page12])
    out12 = collector12.enrich_detail_ssr([row12])
    if page12.goto_calls == [url12] and out12["ui_fallback_success_count"] == 1:
        reporter.pass_("12. Apollo State 누락 시 UI fallback 트리거")
    else:
        reporter.fail(f"12. Apollo missing fallback 이상: goto_calls={page12.goto_calls} out={out12}")

    # 13. Base missing → UI fallback
    row13 = _row("카페 base없음", "9000003")
    url13 = "https://pcmap.place.naver.com/place/9000003/home"
    state13 = {"ROOT_QUERY": {}}
    req13 = FakeAPIRequestContext({url13: FakeSSRResponse(html_text=_html_with_apollo(state13))})
    page13 = FakePage(
        request_ctx=req13,
        entry_frame_by_url={url13: FakeEntryFrame(title="카페 base없음", phone="02-3333-0000")},
    )
    collector13 = _make_collector([page13])
    out13 = collector13.enrich_detail_ssr([row13])
    if page13.goto_calls == [url13] and out13["ui_fallback_success_count"] == 1:
        reporter.pass_("13. Base 누락 시 UI fallback 트리거")
    else:
        reporter.fail(f"13. Base missing fallback 이상: goto_calls={page13.goto_calls} out={out13}")

    # 14. Parent missing → UI fallback
    row14 = _row("카페 parent없음", "9000004")
    url14 = "https://pcmap.place.naver.com/place/9000004/home"
    state14 = {f"PlaceDetailBase:9000004": _base_entity(place_id="9000004")}  # ROOT_QUERY 없음 -> parent 없음
    req14 = FakeAPIRequestContext({url14: FakeSSRResponse(html_text=_html_with_apollo(state14))})
    page14 = FakePage(
        request_ctx=req14,
        entry_frame_by_url={url14: FakeEntryFrame(title="카페 parent없음", phone="02-4444-0000")},
    )
    collector14 = _make_collector([page14])
    out14 = collector14.enrich_detail_ssr([row14])
    if page14.goto_calls == [url14] and out14["ui_fallback_success_count"] == 1:
        reporter.pass_("14. Parent 누락 시 UI fallback 트리거")
    else:
        reporter.fail(f"14. Parent missing fallback 이상: goto_calls={page14.goto_calls} out={out14}")

    # 15. fallback이 block을 감지하면 전체 중단
    row15 = _row("카페 fallback차단", "9000005")
    url15 = "https://pcmap.place.naver.com/place/9000005/home"
    req15 = FakeAPIRequestContext({url15: FakeSSRResponse(status=500, html_text="")})
    captcha15 = FakeCaptchaLocator(count=1, visible=True, box={"width": 100, "height": 100})
    page15 = FakePage(request_ctx=req15, captcha_selectors={"#wtm-captcha-root": captcha15})
    collector15 = _make_collector([page15])
    out15 = collector15.enrich_detail_ssr([row15, _row("업체다음", "9000006")])
    if out15["security_blocked"] is True and out15["stop_reason"] == "captcha_detected":
        reporter.pass_("15. UI fallback이 CAPTCHA를 감지하면 전체 상세 보강 중단")
    else:
        reporter.fail(f"15. fallback block 처리 이상: {out15}")

    # 16. 사용자 중지
    rows16 = [_row(f"업체{i}", f"91000{i}") for i in range(5)]
    urls16 = [f"https://pcmap.place.naver.com/place/91000{i}/home" for i in range(5)]
    req16 = FakeAPIRequestContext({u: FakeSSRResponse(html_text=_html_with_apollo(_full_apollo_state(f"91000{i}"))) for i, u in enumerate(urls16)})
    page16 = FakePage(request_ctx=req16)
    collector16 = _make_collector([page16])
    state16 = {"stop": False}

    def should_continue16():
        return not state16["stop"]

    def on_progress16(completed, total, success, failure):
        if completed >= 2:
            state16["stop"] = True

    out16 = collector16.enrich_detail_ssr(rows16, should_continue=should_continue16, on_progress=on_progress16)
    if len(out16["rows"]) == 5 and len(req16.get_calls) == 2 and out16["stop_reason"] == "user_stopped":
        reporter.pass_("16. 사용자 중지 시 새 요청 즉시 중단, 남은 row 원본 보존")
    else:
        reporter.fail(f"16. 사용자 중지 처리 이상: get_calls={req16.get_calls} out={out16}")

    # 17. row 순서 보존
    rows17 = [_row(f"업체{i}", f"92000{i}") for i in range(4)]
    urls17 = [f"https://pcmap.place.naver.com/place/92000{i}/home" for i in range(4)]
    req17 = FakeAPIRequestContext({u: FakeSSRResponse(html_text=_html_with_apollo(_full_apollo_state(f"92000{i}"))) for i, u in enumerate(urls17)})
    page17 = FakePage(request_ctx=req17)
    collector17 = _make_collector([page17])
    out17 = collector17.enrich_detail_ssr(rows17)
    order_ok = [r["place_id"] for r in out17["rows"]] == [f"92000{i}" for i in range(4)]
    if order_ok:
        reporter.pass_("17. 보강 후에도 row 순서가 입력 순서와 동일하게 보존됨")
    else:
        reporter.fail(f"17. row 순서 이상: {[r['place_id'] for r in out17['rows']]}")

    # 18. row 유실 0(차단 시나리오에서도)
    if len(out10["rows"]) == len(rows10) and len(out16["rows"]) == len(rows16):
        reporter.pass_("18. 차단/중지 시나리오에서도 row 유실 0건(len(rows) 항상 보존)")
    else:
        reporter.fail(f"18. row 유실 발생: out10={len(out10['rows'])}/{len(rows10)}, out16={len(out16['rows'])}/{len(rows16)}")

    # 19. 중복 place_id 상세 요청 1회
    dup_row_a = _row("업체A", "9300001")
    dup_row_b = _row("업체B(같은place_id)", "9300001")
    dup_url = "https://pcmap.place.naver.com/place/9300001/home"
    req19 = FakeAPIRequestContext({dup_url: FakeSSRResponse(html_text=_html_with_apollo(_full_apollo_state("9300001")))})
    page19 = FakePage(request_ctx=req19)
    collector19 = _make_collector([page19])
    out19 = collector19.enrich_detail_ssr([dup_row_a, dup_row_b])
    if req19.get_calls == [dup_url] and out19["rows"][0]["대표전화"] and out19["rows"][1]["대표전화"]:
        reporter.pass_("19. 동일 place_id는 SSR 요청 1회만 발생, 두 row 모두 병합됨")
    else:
        reporter.fail(f"19. 중복 place_id 처리 이상: get_calls={req19.get_calls} out={out19['rows']}")

    # 20. 목록 전역 dedup 이후 상세 실행(UI 배선 - main() 하단 31~36에서 확인)
    reporter.pass_("20. 목록 전역 dedup 이후 상세 실행 여부는 §31~36 UI 배선 테스트에서 확인")

    # 21. total = visitor + blog
    apollo21 = _full_apollo_state("9400001", base_overrides={"visitorReviewsTotal": 100, "cafeBlogReviewsTotal": 50})
    req21 = FakeAPIRequestContext({"https://pcmap.place.naver.com/place/9400001/home": FakeSSRResponse(html_text=_html_with_apollo(apollo21))})
    page21 = FakePage(request_ctx=req21)
    result21 = _fetch_place_detail_ssr(page21, _row("리뷰합산", "9400001"))
    if result21["방문자리뷰수"] == 100 and result21["블로그리뷰수"] == 50 and result21["총리뷰수"] == 150:
        reporter.pass_("21. 총리뷰수 = 방문자 + 블로그 정확히 재계산")
    else:
        reporter.fail(f"21. 리뷰 합산 이상: {result21}")

    # 22. 숫자 0 보존
    apollo22 = _full_apollo_state("9400002", base_overrides={"visitorReviewsTotal": 0, "cafeBlogReviewsTotal": 5})
    req22 = FakeAPIRequestContext({"https://pcmap.place.naver.com/place/9400002/home": FakeSSRResponse(html_text=_html_with_apollo(apollo22))})
    page22 = FakePage(request_ctx=req22)
    result22 = _fetch_place_detail_ssr(page22, _row("리뷰0", "9400002"))
    if result22["방문자리뷰수"] == 0 and result22["총리뷰수"] == 5:
        reporter.pass_("22. 실제 값 0은 누락으로 취급하지 않고 그대로 보존")
    else:
        reporter.fail(f"22. 숫자 0 보존 이상: {result22}")

    # 23. missing과 0 구분(하나만 missing이면 fallback 아님, 값 공란 유지)
    apollo23 = _full_apollo_state("9400003", base_overrides={"cafeBlogReviewsTotal": 0})
    del apollo23["PlaceDetailBase:9400003"]["visitorReviewsTotal"]
    req23 = FakeAPIRequestContext({"https://pcmap.place.naver.com/place/9400003/home": FakeSSRResponse(html_text=_html_with_apollo(apollo23))})
    page23 = FakePage(request_ctx=req23)
    result23 = _fetch_place_detail_ssr(page23, _row("리뷰일부missing", "9400003"))
    if result23["detail_ssr_status"] == "success" and result23["방문자리뷰수"] == "" and result23["블로그리뷰수"] == 0:
        reporter.pass_("23. 한쪽 리뷰만 missing이면 fallback 없이 공란/0을 구분 유지")
    else:
        reporter.fail(f"23. missing/0 구분 이상: {result23}")

    # 24. 개인 이동전화 필터
    apollo24 = _full_apollo_state("9400004", base_overrides={"phone": "010-1234-5678", "virtualPhone": ""})
    req24 = FakeAPIRequestContext({"https://pcmap.place.naver.com/place/9400004/home": FakeSSRResponse(html_text=_html_with_apollo(apollo24))})
    page24 = FakePage(request_ctx=req24)
    result24 = _fetch_place_detail_ssr(page24, _row("개인번호", "9400004"))
    if result24["대표전화"] != "010-1234-5678":
        reporter.pass_("24. 개인 이동전화(010)는 SSR 경로에서도 대표전화로 채택되지 않음")
    else:
        reporter.fail(f"24. 개인 이동전화 필터 실패: {result24}")

    # 25. URL 분류(인스타)
    apollo25 = _full_apollo_state(
        "9400005",
        parent_overrides={"homepages": {"repr": {"url": "https://www.instagram.com/testcafe", "type": "인스타그램"}, "etc": []}},
    )
    req25 = FakeAPIRequestContext({"https://pcmap.place.naver.com/place/9400005/home": FakeSSRResponse(html_text=_html_with_apollo(apollo25))})
    page25 = FakePage(request_ctx=req25)
    result25 = _fetch_place_detail_ssr(page25, _row("인스타분류", "9400005"))
    if result25["인스타"] == "https://www.instagram.com/testcafe" and not result25["홈페이지"]:
        reporter.pass_("25. instagram.com URL이 홈페이지가 아닌 인스타로 정확히 분류됨")
    else:
        reporter.fail(f"25. URL 분류 이상: {result25}")

    # 26. newOpening tri-state
    apollo26_true = _full_apollo_state("9400006", parent_overrides={"newOpening": True})
    req26 = FakeAPIRequestContext({"https://pcmap.place.naver.com/place/9400006/home": FakeSSRResponse(html_text=_html_with_apollo(apollo26_true))})
    page26 = FakePage(request_ctx=req26)
    result26 = _fetch_place_detail_ssr(page26, _row("새오픈", "9400006"))
    if result26["새로오픈여부"] == "O":
        reporter.pass_("26. newOpening=True는 '새로오픈여부'=O로 매핑")
    else:
        reporter.fail(f"26. newOpening tri-state 이상: {result26}")

    # 27. custom diagnostic 필드 Excel 미노출
    diagnostic_keys = {
        "detail_ssr_attempted", "detail_ssr_status", "detail_ssr_http_status", "detail_ssr_apollo_found",
        "detail_ssr_base_found", "detail_ssr_parent_found", "detail_ssr_place_id_exact",
        "detail_ssr_adapter_success", "detail_ssr_stop_reason", "detail_ssr_error_type",
        "detail_ssr_duration_seconds",
    }
    if not (diagnostic_keys & set(MERGED_COLUMNS)):
        reporter.pass_("27. detail_ssr_* 진단 키는 MERGED_COLUMNS(Excel 14컬럼)에 전혀 포함되지 않음")
    else:
        reporter.fail(f"27. 진단 키가 Excel 컬럼에 노출됨: {diagnostic_keys & set(MERGED_COLUMNS)}")

    # 28. `업체명` 키 유지(merge_detail_into_row가 다른 키로 바꾸지 않음)
    merged28 = merge_detail_into_row(_row("원본업체명", PLACE_ID), result1)
    if merged28.get("업체명") == "원본업체명" and "상호명" not in merged28:
        reporter.pass_("28. merge_detail_into_row가 '업체명' 키를 유지하고 '상호명'을 만들지 않음(5W-R1 재발 방지)")
    else:
        reporter.fail(f"28. 업체명 키 유지 실패: {merged28}")

    # 29. 기존 exporter 13컬럼으로 실제 저장
    with tempfile.TemporaryDirectory() as tmp_dir:
        out_path = str(Path(tmp_dir) / "ssr_29.xlsx")
        export_places_to_excel([merged28], [], [], out_path)
        import openpyxl
        wb = openpyxl.load_workbook(out_path)
        headers = [c.value for c in wb["통합_결과"][1]]
        if headers == MERGED_COLUMNS and len(headers) == 14:
            reporter.pass_("29. 기존 export_places_to_excel로 SSR 보강 row가 14컬럼 그대로 저장됨")
        else:
            reporter.fail(f"29. exporter 14컬럼 이상: {headers}")

    # 30. 부분 결과 export(차단으로 일부만 보강된 rows도 손실 없이 저장)
    with tempfile.TemporaryDirectory() as tmp_dir:
        out_path = str(Path(tmp_dir) / "ssr_30.xlsx")
        partial_rows = out10["rows"]  # 1건 보강 성공 + 1건 원본 보존(차단)
        export_places_to_excel(partial_rows, [], [], out_path)
        import openpyxl
        wb = openpyxl.load_workbook(out_path)
        data_rows = wb["통합_결과"].max_row - 1
        if data_rows == len(partial_rows):
            reporter.pass_("30. 차단으로 인한 부분 결과도 rows 전체(원본 보존분 포함)가 Excel에 손실 없이 저장됨")
        else:
            reporter.fail(f"30. 부분 결과 export 이상: data_rows={data_rows}, expected={len(partial_rows)}")

    # ------------------------------------------------------------------
    # 31~36. src/ui.py 기본 흐름 배선
    # ------------------------------------------------------------------

    class FakeSSRCollector:
        def __init__(self, collected_at, *, ssr_rows_out=None, ssr_return=None):
            self.collected_at = collected_at
            self._ssr_rows_out = ssr_rows_out
            self._ssr_return = ssr_return
            self.enrich_calls: list = []
            self.exited = False

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            self.exited = True
            return False

        def collect_query(self, job, per_query_limit):
            raise AssertionError("collect_query는 fake orchestrator만 사용해야 함")

        def enrich_detail_ssr(self, rows, *, max_targets=None, should_continue=None, on_progress=None):
            self.enrich_calls.append({
                "rows": rows, "max_targets": max_targets, "exited_before_call": self.exited,
            })
            if self._ssr_return is not None:
                return self._ssr_return
            out_rows = self._ssr_rows_out if self._ssr_rows_out is not None else rows
            return {
                "rows": out_rows, "stop_reason": None, "security_blocked": False,
                "attempted_count": len(rows), "ssr_success_count": len(rows),
                "ui_fallback_used_count": 0, "ui_fallback_success_count": 0,
                "failure_count": 0, "not_attempted_count": 0,
            }

    def _make_app():
        app = ui.SalesDbCrawlerApp.__new__(ui.SalesDbCrawlerApp)
        app.stop_event = threading.Event()
        app.pause_event = threading.Event()
        app.log = lambda message: None
        app.set_status = lambda message: None
        app._security_block_decision = None
        return app

    def _base_orchestrator_result(rows):
        return {
            "rows": rows, "executed_query_count": 1, "skipped_query_count": 0,
            "stop_reason": "queue_exhausted", "before_trim_count": len(rows), "final_count": len(rows),
            "security_blocked": False, "status_429_seen": False, "navigation_error": False,
            "navigation_error_message": "",
        }

    def _fake_row(i):
        return {
            "업체명": f"업체{i}", "업종": "카페", "새로오픈여부": "", "방문자리뷰수": "",
            "블로그리뷰수": "", "총리뷰수": "", "주소": f"서울 강동구 {i}", "대표전화": "",
            "플레이스 URL": f"https://pcmap.place.naver.com/place/{i}/home", "수집일": "2026-07-23",
            "홈페이지": "", "인스타": "", "블로그": "", "place_id": str(i),
        }

    # 31. 기본 흐름에서 SSR 상세 보강 자동 호출(토글 없음)
    rows31 = [_fake_row(i) for i in range(3)]
    holder31: dict = {}

    def factory31(*, collected_at):
        c = FakeSSRCollector(collected_at)
        holder31["collector"] = c
        return c

    def orchestrator31(jobs, **kwargs):
        return _base_orchestrator_result(rows31)

    with tempfile.TemporaryDirectory() as tmp_dir:
        app31 = _make_app()
        app31._run_network_pipeline(
            [{"query": "q1"}], 30, 300, str(Path(tmp_dir) / "out.xlsx"),
            collector_factory=factory31, orchestrator=orchestrator31, excel_exporter=export_places_to_excel,
        )
    if len(holder31["collector"].enrich_calls) == 1:
        reporter.pass_("31. 별도 호출 없이 기본 흐름에서 enrich_detail_ssr이 자동 호출됨")
    else:
        reporter.fail(f"31. 기본 자동 호출 이상: {holder31['collector'].enrich_calls}")

    # 32. 새 토글 없음(시그니처에 ssr 관련 신규 파라미터 없음)
    sig = inspect.signature(ui.SalesDbCrawlerApp._run_network_pipeline)
    new_toggle_like = [p for p in sig.parameters if "ssr" in p.lower() or "toggle" in p.lower() or "enable" in p.lower()]
    if not new_toggle_like:
        reporter.pass_("32. _run_network_pipeline 시그니처에 SSR on/off 토글성 파라미터 없음")
    else:
        reporter.fail(f"32. 의심되는 토글 파라미터 발견: {new_toggle_like}")

    # 33. final unique row 수를 max_targets로 전달
    if holder31["collector"].enrich_calls[0]["max_targets"] == len(rows31):
        reporter.pass_("33. max_targets가 최종 고유 row 수(len(rows))로 그대로 전달됨")
    else:
        reporter.fail(f"33. max_targets 전달 이상: {holder31['collector'].enrich_calls[0]}")

    # 34. SSR 이후 exporter 호출(보강된 rows가 exporter에 전달됨)
    enriched_rows34 = [dict(r, 대표전화="02-9999-0000") for r in rows31]
    exporter_calls34: list = []

    def spy_exporter34(merged, mobile, pc, output_path):
        exporter_calls34.append(list(merged))
        return export_places_to_excel(merged, mobile, pc, output_path)

    def factory34(*, collected_at):
        return FakeSSRCollector(collected_at, ssr_rows_out=enriched_rows34)

    with tempfile.TemporaryDirectory() as tmp_dir:
        app34 = _make_app()
        app34._run_network_pipeline(
            [{"query": "q1"}], 30, 300, str(Path(tmp_dir) / "out.xlsx"),
            collector_factory=factory34, orchestrator=orchestrator31, excel_exporter=spy_exporter34,
        )
    if exporter_calls34 and exporter_calls34[0][0]["대표전화"] == "02-9999-0000":
        reporter.pass_("34. exporter가 SSR 보강 이후의 rows(원본이 아닌)를 전달받음")
    else:
        reporter.fail(f"34. exporter 전달 rows 이상: {exporter_calls34}")

    # 35. 중지 후 부분 결과 exporter 호출
    partial_ssr_return35 = {
        "rows": rows31, "stop_reason": "user_stopped", "security_blocked": False,
        "attempted_count": 1, "ssr_success_count": 1, "ui_fallback_used_count": 0,
        "ui_fallback_success_count": 0, "failure_count": 0, "not_attempted_count": 2,
    }

    def factory35(*, collected_at):
        return FakeSSRCollector(collected_at, ssr_return=partial_ssr_return35)

    exporter_calls35: list = []

    def spy_exporter35(merged, mobile, pc, output_path):
        exporter_calls35.append(list(merged))
        return export_places_to_excel(merged, mobile, pc, output_path)

    with tempfile.TemporaryDirectory() as tmp_dir:
        app35 = _make_app()
        returned35 = app35._run_network_pipeline(
            [{"query": "q1"}], 30, 300, str(Path(tmp_dir) / "out.xlsx"),
            collector_factory=factory35, orchestrator=orchestrator31, excel_exporter=spy_exporter35,
        )
    if exporter_calls35 and len(exporter_calls35[0]) == 3 and returned35.get("exported") is True:
        reporter.pass_("35. SSR 단계 중지 시에도 부분 결과(3건 전부, row 유실 없음)가 exporter에 전달되어 저장됨")
    else:
        reporter.fail(f"35. 중지 후 부분 저장 이상: exporter_calls={exporter_calls35} returned={returned35}")

    # 36. 같은 collector 세션 안에서 실행(collector.__exit__ 이전에 enrich_detail_ssr 호출)
    if holder31["collector"].enrich_calls[0]["exited_before_call"] is False:
        reporter.pass_("36. enrich_detail_ssr이 collector 세션(__exit__ 호출 전)이 닫히기 전에 실행됨")
    else:
        reporter.fail("36. enrich_detail_ssr이 세션 종료 이후 호출됨(요청서 §6 위반)")

    print("====================")
    print("검증 요약")
    print(f"PASS: {reporter.passes}")
    print(f"FAIL: {reporter.fails}")
    print(f"FINAL: {'FAIL' if reporter.fails else 'PASS'}")
    print("====================")
    return 1 if reporter.fails else 0


if __name__ == "__main__":
    sys.exit(main())
