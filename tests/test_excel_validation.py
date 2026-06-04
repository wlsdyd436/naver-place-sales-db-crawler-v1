from pathlib import Path
import re
import sys

import pandas as pd
from openpyxl import load_workbook


# 2026-06-05: V1.2 최종 납품 Excel 3시트 무결성 검증 스크립트입니다.
DEFAULT_EXCEL_PATTERN = "naver_place_merged_db*.xlsx"
OUTPUT_DIR = Path("output")

REQUIRED_SHEETS = ["통합_결과", "원본_모바일", "원본_PC"]

EXPECTED_COLUMNS = {
    "통합_결과": [
        "업체명",
        "업종",
        "새로오픈여부",
        "리뷰수",
        "주소",
        "대표전화",
        "플레이스 URL",
        "수집일",
    ],
    "원본_모바일": [
        "업체명",
        "업종",
        "주소",
        "대표전화",
        "플레이스 URL",
        "수집일",
    ],
    # 2026-06-05: 현재 exporter.py는 원본_PC에 주소를 포함해 저장합니다.
    "원본_PC": [
        "업체명",
        "업종",
        "새로오픈여부",
        "리뷰수",
        "주소",
        "수집일",
    ],
}


class ValidationReporter:
    def __init__(self):
        self.pass_count = 0
        self.fail_count = 0
        self.warn_count = 0

    def pass_(self, message: str) -> None:
        self.pass_count += 1
        print(f"[PASS] {message}")

    def fail(self, message: str) -> None:
        self.fail_count += 1
        print(f"[FAIL] {message}")

    def warn(self, message: str) -> None:
        self.warn_count += 1
        print(f"[WARN] {message}")

    def summary(self) -> None:
        final = "FAIL" if self.fail_count else "PASS"
        print("====================")
        print("검증 요약")
        print(f"PASS: {self.pass_count}")
        print(f"FAIL: {self.fail_count}")
        print(f"WARN: {self.warn_count}")
        print(f"FINAL: {final}")
        print("====================")


def find_target_excel() -> Path:
    if len(sys.argv) >= 2:
        return Path(sys.argv[1])

    candidates = sorted(
        OUTPUT_DIR.glob(DEFAULT_EXCEL_PATTERN),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else OUTPUT_DIR / "naver_place_merged_db.xlsx"


def is_blank(value) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return str(value).strip() == ""


def normalize_cell(value) -> str:
    if is_blank(value):
        return ""
    return str(value).strip()


def read_excel_sheets(path: Path) -> tuple[list[str], dict[str, pd.DataFrame]]:
    # 2026-06-05: pandas 우선, 실패 시 openpyxl로 fallback합니다.
    try:
        excel_file = pd.ExcelFile(path)
        frames = {
            sheet_name: pd.read_excel(path, sheet_name=sheet_name, dtype=str)
            for sheet_name in excel_file.sheet_names
        }
        return excel_file.sheet_names, frames
    except Exception:
        workbook = load_workbook(path, data_only=True)
        frames = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(worksheet.values)
            if not rows:
                frames[sheet_name] = pd.DataFrame()
                continue
            headers = list(rows[0])
            frames[sheet_name] = pd.DataFrame(rows[1:], columns=headers)
        return workbook.sheetnames, frames


def validate_sheet_names(reporter: ValidationReporter, sheet_names: list[str]) -> None:
    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in sheet_names]
    if missing:
        reporter.fail(f"필수 시트 누락: {missing}")
    else:
        reporter.pass_("필수 시트 확인")

    extra = [sheet for sheet in sheet_names if sheet not in REQUIRED_SHEETS]
    if extra:
        reporter.warn(f"추가 시트 발견: {extra}")


def validate_columns(
    reporter: ValidationReporter, frames: dict[str, pd.DataFrame]
) -> None:
    for sheet_name, expected_columns in EXPECTED_COLUMNS.items():
        frame = frames.get(sheet_name)
        if frame is None:
            continue
        actual_columns = list(frame.columns)
        if actual_columns == expected_columns:
            reporter.pass_(f"{sheet_name} 컬럼 순서 확인")
        else:
            reporter.fail(
                f"{sheet_name} 컬럼 불일치 - 기대: {expected_columns}, 실제: {actual_columns}"
            )


def validate_required_values(
    reporter: ValidationReporter, frame: pd.DataFrame, sheet_name: str
) -> None:
    failed = False
    for column in ["업체명", "수집일"]:
        if column not in frame.columns:
            reporter.fail(f"{sheet_name} 필수 컬럼 없음: {column}")
            failed = True
            continue
        for index, value in frame[column].items():
            if is_blank(value):
                reporter.fail(f"{sheet_name} 필수값 누락 - {index + 2}행 {column}")
                failed = True
    if not failed:
        reporter.pass_(f"{sheet_name} 필수값 확인")


def validate_new_open(reporter: ValidationReporter, frame: pd.DataFrame) -> None:
    failed = False
    for index, value in frame["새로오픈여부"].items():
        text = normalize_cell(value)
        if text not in ("", "O"):
            reporter.fail(f"새로오픈여부 값 오류 - {index + 2}행 값: {text}")
            failed = True
    if not failed:
        reporter.pass_("새로오픈여부 값 확인")


def validate_review_count(reporter: ValidationReporter, frame: pd.DataFrame) -> None:
    failed = False
    for index, value in frame["리뷰수"].items():
        text = normalize_cell(value)
        if not text:
            continue
        if not text.replace(",", "").isdigit():
            reporter.fail(f"리뷰수 형식 오류 - {index + 2}행 값: {text}")
            failed = True
    if not failed:
        reporter.pass_("리뷰수 형식 확인")


def validate_phone(reporter: ValidationReporter, frame: pd.DataFrame) -> None:
    failed = False
    phone_pattern = re.compile(r"^[0-9-]+$")
    for index, value in frame["대표전화"].items():
        text = normalize_cell(value)
        if not text:
            continue
        if not phone_pattern.fullmatch(text):
            reporter.fail(f"대표전화 형식 오류 - {index + 2}행 값: {text}")
            failed = True
    if not failed:
        reporter.pass_("대표전화 형식 확인")


def validate_row_count(
    reporter: ValidationReporter, merged_frame: pd.DataFrame, mobile_frame: pd.DataFrame
) -> None:
    if len(merged_frame) == len(mobile_frame):
        reporter.pass_("병합 행 개수 확인")
    else:
        reporter.fail(
            f"병합 행 개수 불일치 - 통합_결과: {len(merged_frame)}, 원본_모바일: {len(mobile_frame)}"
        )


def validate_empty_data_warning(
    reporter: ValidationReporter, frames: dict[str, pd.DataFrame]
) -> None:
    for sheet_name in REQUIRED_SHEETS:
        frame = frames.get(sheet_name)
        if frame is not None and frame.empty:
            reporter.warn(f"{sheet_name} 데이터 0건")


def main() -> int:
    reporter = ValidationReporter()
    excel_path = find_target_excel()

    if excel_path.exists():
        reporter.pass_(f"엑셀 파일 존재 확인: {excel_path}")
    else:
        reporter.fail(f"엑셀 파일 없음: {excel_path}")
        reporter.summary()
        return 1

    try:
        sheet_names, frames = read_excel_sheets(excel_path)
        reporter.pass_("엑셀 파일 읽기 확인")
    except Exception as exc:
        reporter.fail(f"엑셀 파일 읽기 실패: {exc}")
        reporter.summary()
        return 1

    validate_sheet_names(reporter, sheet_names)
    validate_columns(reporter, frames)
    validate_empty_data_warning(reporter, frames)

    if all(sheet in frames for sheet in REQUIRED_SHEETS):
        merged_frame = frames["통합_결과"]
        mobile_frame = frames["원본_모바일"]
        pc_frame = frames["원본_PC"]

        validate_required_values(reporter, merged_frame, "통합_결과")
        validate_required_values(reporter, mobile_frame, "원본_모바일")
        validate_required_values(reporter, pc_frame, "원본_PC")
        validate_new_open(reporter, merged_frame)
        validate_review_count(reporter, merged_frame)
        validate_phone(reporter, merged_frame)
        validate_row_count(reporter, merged_frame, mobile_frame)

    reporter.summary()
    return 1 if reporter.fail_count else 0


if __name__ == "__main__":
    sys.exit(main())
