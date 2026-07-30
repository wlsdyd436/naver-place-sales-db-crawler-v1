from pathlib import Path
import sys
import tempfile

from openpyxl import load_workbook


# 5M-R1(2026-07-23): exporter 통합_결과 스키마를 13컬럼으로 확정(방문자/블로그/
# 총리뷰수 분리, 기존 단일 "리뷰수" 컬럼 제거). 원본_모바일/원본_PC(MOBILE_COLUMNS/
# PC_COLUMNS)는 legacy 엔진 전용이라 이번 변경 대상이 아니며 그대로 검증한다.
ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.exporter import (
    MERGED_COLUMNS,
    MOBILE_COLUMNS,
    PC_COLUMNS,
    export_places_to_excel,
)


EXPECTED_MERGED = [
    "업체명",
    "업종",
    "새로오픈여부",
    "방문자리뷰수",
    "블로그리뷰수",
    "총리뷰수",
    "주소",
    "대표전화",
    "플레이스 URL",
    "수집일",
    "홈페이지",
    "인스타",
    "블로그",
    "추가 링크",
]
EXPECTED_MOBILE = ["업체명", "업종", "주소", "대표전화", "플레이스 URL", "수집일"]
EXPECTED_PC = ["업체명", "업종", "새로오픈여부", "리뷰수", "주소", "수집일"]


class ValidationReporter:
    def __init__(self):
        self.pass_count = 0
        self.fail_count = 0
        self.warn_count = 0

    def pass_(self, message: str) -> None:
        self.pass_count += 1
        print(f"[PASS] {message}")

    def fail(self, message: str) -> None:
        self.fail_count += 1
        print(f"[FAIL] {message}")

    def summary(self) -> None:
        final = "FAIL" if self.fail_count else "PASS"
        print("====================")
        print("검증 요약")
        print(f"PASS: {self.pass_count}")
        print(f"FAIL: {self.fail_count}")
        print(f"WARN: {self.warn_count}")
        print(f"FINAL: {final}")
        print("====================")


def _new_engine_row():
    # collect_pc_full 결과 형태(가산 필드 + place_id 포함). 5M-R1: 방문자/블로그
    # 리뷰수는 정수로, 총리뷰수는 둘 다 확인됐을 때만 합산해 저장한다.
    return {
        "업체명": "오베르캄프 본점",
        "업종": "베이커리",
        "새로오픈여부": "",
        "방문자리뷰수": 2400,
        "블로그리뷰수": 69,
        "총리뷰수": 2469,
        "주소": "서울 강동구 성내로14길 48 1층",
        "대표전화": "0507-1387-4967",
        "플레이스 URL": "https://pcmap.place.naver.com/restaurant/1171815551/home",
        "place_id": "1171815551",
        "수집일": "2026-07-06",
        "홈페이지": "",
        "인스타": "https://www.instagram.com/oberkampf.kr",
        "블로그": "",
    }


def _legacy_row():
    # SNS 필드가 없는 기존 흐름 행(모바일/merge 결과 형태). 방문자/블로그 리뷰수는
    # 미확인(빈 문자열)이므로 총리뷰수도 공란이어야 한다(부분합계 금지).
    return {
        "업체명": "레거시 카페",
        "업종": "카페",
        "새로오픈여부": "O",
        "방문자리뷰수": "",
        "블로그리뷰수": "",
        "총리뷰수": "",
        "주소": "서울 강동구 어딘가",
        "대표전화": "02-000-0000",
        "플레이스 URL": "https://m.place.naver.com/place/1/home",
        "수집일": "2026-07-06",
    }


def _read_headers(path, sheet_name):
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name]
    return [cell.value for cell in worksheet[1]]


def _read_rows(path, sheet_name):
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name]
    headers = [cell.value for cell in worksheet[1]]
    rows = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, values)))
    return rows


# ---------------------------------------------------------------------------
# 1. 모듈 상수 스키마
# ---------------------------------------------------------------------------


def check_merged_columns_constant(reporter: ValidationReporter) -> None:
    if MERGED_COLUMNS == EXPECTED_MERGED:
        reporter.pass_("MERGED_COLUMNS = PAGE300-6E-V3 확정 14컬럼(방문자/블로그/총리뷰수 분리 + 추가 링크 포함)")
    else:
        reporter.fail(f"MERGED_COLUMNS이 예상 14컬럼과 다름: {MERGED_COLUMNS}")

    if "리뷰수" not in MERGED_COLUMNS:
        reporter.pass_("기존 단일 '리뷰수' 컬럼이 통합_결과에서 제거됨")
    else:
        reporter.fail("기존 단일 '리뷰수' 컬럼이 여전히 남아있음")


def check_other_columns_unchanged(reporter: ValidationReporter) -> None:
    if MOBILE_COLUMNS == EXPECTED_MOBILE and PC_COLUMNS == EXPECTED_PC:
        reporter.pass_("MOBILE_COLUMNS / PC_COLUMNS 무변경")
    else:
        reporter.fail(f"원본 시트 컬럼이 변경됨: mobile={MOBILE_COLUMNS}, pc={PC_COLUMNS}")


# ---------------------------------------------------------------------------
# 2. 실제 저장 파일 스키마
# ---------------------------------------------------------------------------


def check_exported_merged_headers(reporter: ValidationReporter) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "schema_test.xlsx"
        export_places_to_excel([_new_engine_row(), _legacy_row()], [], [], str(path))

        merged_headers = _read_headers(path, "통합_결과")
        checks = {
            "통합_결과 헤더 14개(PAGE300-6E-V3 확정 스키마)": merged_headers == EXPECTED_MERGED,
            "place_id는 통합_결과 컬럼에 미노출": "place_id" not in merged_headers,
            "원본_모바일 헤더 무변경": _read_headers(path, "원본_모바일") == EXPECTED_MOBILE,
            "원본_PC 헤더 무변경": _read_headers(path, "원본_PC") == EXPECTED_PC,
        }
        if all(checks.values()):
            reporter.pass_("저장 파일 3시트 헤더: 통합_결과는 14컬럼, place_id 미노출, 원본 시트 불변")
        else:
            failed = [name for name, ok in checks.items() if not ok]
            reporter.fail(f"저장 헤더 검증 실패 항목: {failed} (통합_결과={merged_headers})")


def check_new_engine_row_values(reporter: ValidationReporter) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "values_test.xlsx"
        export_places_to_excel([_new_engine_row()], [], [], str(path))
        rows = _read_rows(path, "통합_결과")
        row = rows[0] if rows else {}
        ok = (
            row.get("업체명") == "오베르캄프 본점"
            and row.get("대표전화") == "0507-1387-4967"
            and row.get("플레이스 URL") == "https://pcmap.place.naver.com/restaurant/1171815551/home"
            and row.get("인스타") == "https://www.instagram.com/oberkampf.kr"
            and row.get("방문자리뷰수") == 2400
            and row.get("블로그리뷰수") == 69
            and row.get("총리뷰수") == 2469
        )
        if ok:
            reporter.pass_("PC full 엔진 row가 통합_결과 13컬럼에 정확히 매핑(대표전화/URL/인스타/리뷰 분리 포함)")
        else:
            reporter.fail(f"새 엔진 row 매핑 결과가 예상과 다름: {row}")


def check_review_columns_numeric_cells(reporter: ValidationReporter) -> None:
    # 방문자/블로그/총리뷰수는 Excel 수식이 아닌 검증된 정수 셀로 저장돼야 하고,
    # 미확인 값은 공란(빈 셀)이어야 한다(부분 합계로 채우지 않음).
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "numeric_test.xlsx"
        export_places_to_excel([_new_engine_row(), _legacy_row()], [], [], str(path))
        rows = _read_rows(path, "통합_결과")
        new_engine_row = rows[0] if rows else {}
        legacy_row = rows[1] if len(rows) > 1 else {}
        ok = (
            isinstance(new_engine_row.get("방문자리뷰수"), int)
            and isinstance(new_engine_row.get("블로그리뷰수"), int)
            and isinstance(new_engine_row.get("총리뷰수"), int)
            and new_engine_row.get("총리뷰수") == 2469
            and (legacy_row.get("방문자리뷰수") is None or str(legacy_row.get("방문자리뷰수")).strip() == "")
            and (legacy_row.get("총리뷰수") is None or str(legacy_row.get("총리뷰수")).strip() == "")
        )
        if ok:
            reporter.pass_("방문자/블로그/총리뷰수가 숫자 셀로 저장되고, 미확인 값은 공란 유지")
        else:
            reporter.fail(f"리뷰수 숫자 셀/공란 처리 결과가 예상과 다름: new={new_engine_row}, legacy={legacy_row}")


def check_legacy_row_blank_sns(reporter: ValidationReporter) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "legacy_test.xlsx"
        export_places_to_excel([_legacy_row()], [], [], str(path))
        rows = _read_rows(path, "통합_결과")
        row = rows[0] if rows else {}

        def _blank(value):
            return value is None or str(value).strip() == ""

        ok = (
            row.get("업체명") == "레거시 카페"
            and _blank(row.get("홈페이지"))
            and _blank(row.get("인스타"))
            and _blank(row.get("블로그"))
        )
        if ok:
            reporter.pass_("SNS 필드 없는 기존 흐름 행은 홈페이지/인스타/블로그가 빈칸(get 기본값)")
        else:
            reporter.fail(f"레거시 행 빈칸 처리 결과가 예상과 다름: {row}")


def check_fourteen_columns_extra_links_last(reporter: ValidationReporter) -> None:
    """PAGE300-6E-V3: 통합_결과는 정확히 14컬럼이고 "추가 링크"가 마지막이어야 한다."""
    if len(MERGED_COLUMNS) == 14 and MERGED_COLUMNS[-1] == "추가 링크" and MERGED_COLUMNS[:13] == EXPECTED_MERGED[:13]:
        reporter.pass_("MERGED_COLUMNS이 정확히 14컬럼이고 '추가 링크'가 블로그 다음 마지막 컬럼")
    else:
        reporter.fail(f"14컬럼/추가 링크 마지막 배치 검증 실패: {MERGED_COLUMNS}")


def check_legacy_thirteen_column_row_backward_compatible(reporter: ValidationReporter) -> None:
    """"추가 링크" 키 자체가 없는 기존 13컬럼 입력 row도 예외 없이 공란으로 export되어야 한다."""
    legacy_row_without_extra = _legacy_row()
    if "추가 링크" in legacy_row_without_extra:
        reporter.fail("픽스처 오류: _legacy_row에 추가 링크 키가 이미 존재함")
        return
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "legacy13_test.xlsx"
        export_places_to_excel([legacy_row_without_extra], [], [], str(path))
        rows = _read_rows(path, "통합_결과")
        row = rows[0] if rows else {}
        value = row.get("추가 링크")
        if value is None or str(value).strip() == "":
            reporter.pass_("추가 링크 키가 없는 기존 13컬럼 row도 예외 없이 공란으로 export됨(하위 호환)")
        else:
            reporter.fail(f"13컬럼 하위 호환 export 결과가 예상과 다름: {value!r}")


def check_extra_links_multi_line_value_and_wrap_text(reporter: ValidationReporter) -> None:
    """여러 URL이 줄바꿈으로 저장된 "추가 링크" 값이 셀에 그대로 보존되고
    wrap_text + 상단 정렬이 적용되는지 확인한다."""
    row = _new_engine_row()
    row["추가 링크"] = "[카카오채널] https://pf.kakao.com/_example\n[유튜브] https://youtube.com/@example"
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "extra_links_test.xlsx"
        export_places_to_excel([row], [], [], str(path))
        workbook = load_workbook(path, data_only=True)
        worksheet = workbook["통합_결과"]
        header = [cell.value for cell in worksheet[1]]
        col_index = header.index("추가 링크") + 1
        data_cell = worksheet.cell(row=2, column=col_index)
        ok = (
            data_cell.value == "[카카오채널] https://pf.kakao.com/_example\n[유튜브] https://youtube.com/@example"
            and data_cell.alignment.wrap_text is True
            and data_cell.alignment.vertical == "top"
        )
        if ok:
            reporter.pass_("여러 URL 줄바꿈 값이 셀에 그대로 보존되고 wrap_text+상단 정렬 적용됨")
        else:
            reporter.fail(f"추가 링크 셀 값/서식 검증 실패: value={data_cell.value!r}, alignment={data_cell.alignment}")


def check_signature_unchanged(reporter: ValidationReporter) -> None:
    # export_places_to_excel 시그니처(위치 인자 4개)와 반환 경로 계약 유지 확인.
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "sig_test.xlsx"
        returned = export_places_to_excel([_legacy_row()], [_legacy_row()], [_legacy_row()], str(path))
        if returned == str(path) and Path(returned).exists():
            reporter.pass_("export_places_to_excel 시그니처/반환 경로 계약 유지")
        else:
            reporter.fail(f"export 반환/저장 계약이 예상과 다름: {returned!r}")


def main() -> int:
    reporter = ValidationReporter()

    check_merged_columns_constant(reporter)
    check_other_columns_unchanged(reporter)
    check_exported_merged_headers(reporter)
    check_new_engine_row_values(reporter)
    check_review_columns_numeric_cells(reporter)
    check_legacy_row_blank_sns(reporter)
    check_fourteen_columns_extra_links_last(reporter)
    check_legacy_thirteen_column_row_backward_compatible(reporter)
    check_extra_links_multi_line_value_and_wrap_text(reporter)
    check_signature_unchanged(reporter)

    reporter.summary()
    return 1 if reporter.fail_count else 0


def test_standalone_suite():
    assert main() == 0


if __name__ == "__main__":
    sys.exit(main())
