from pathlib import Path
import sys
import tempfile
import threading
from types import SimpleNamespace

import openpyxl


# ARCH-300C WIRE-2C-1: src.ui._run_network_pipeline이 실제 export_places_to_excel을
# 임시 .xlsx 경로에 호출했을 때의 산출물을 검증하는 standalone 스크립트.
# 브라우저/collector/orchestrator는 여전히 fake이며(live 없음), 저장 단계만
# 진짜 exporter를 실행한다(실제 파일 IO는 tempfile.TemporaryDirectory 안에서만
# 발생하고 테스트 종료 시 자동 삭제된다).
ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src import ui, ui_home_stage
from src.exporter import MERGED_COLUMNS, export_places_to_excel
from src.diagnostics import DiagnosticArtifact


class ValidationReporter:
    def __init__(self):
        self.pass_count = 0
        self.fail_count = 0
        self.warn_count = 0

    def pass_(self, message: str) -> None:
        self.pass_count += 1
        print(f"[PASS] {message}")

    def fail(self, message: str) -> None:
        self.fail_count += 1
        print(f"[FAIL] {message}")

    def summary(self) -> None:
        final = "FAIL" if self.fail_count else "PASS"
        print("====================")
        print("검증 요약")
        print(f"PASS: {self.pass_count}")
        print(f"FAIL: {self.fail_count}")
        print(f"WARN: {self.warn_count}")
        print(f"FINAL: {final}")
        print("====================")


def _make_app():
    app = ui.SalesDbCrawlerApp.__new__(ui.SalesDbCrawlerApp)
    app.stop_event = threading.Event()
    app.pause_event = threading.Event()
    app.log = lambda message: None
    app.set_status = lambda message: None
    app.after = lambda delay, func, *args: func(*args)
    app.duplicate_removed_var = SimpleNamespace(set=lambda value: None)
    app._security_block_decision = None
    return app


class FakeNetworkCollector:
    def __init__(self, collected_at):
        self.collected_at = collected_at

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False

    def collect_query(self, job, per_query_limit):
        raise AssertionError("이 통합 테스트에서는 collect_query가 직접 호출되면 안 됨(fake orchestrator만 사용)")


def _fake_collector_factory(*, collected_at, pause_event=None, stop_event=None):
    return FakeNetworkCollector(collected_at)


def _make_fake_orchestrator(result: dict):
    def fake_orchestrator(jobs, **kwargs):
        return result

    return fake_orchestrator


def _base_result(**overrides) -> dict:
    result = {
        "rows": [],
        "executed_query_count": 1,
        "skipped_query_count": 0,
        "stop_reason": "queue_exhausted",
        "before_trim_count": 0,
        "final_count": 0,
        "security_blocked": False,
        "status_429_seen": False,
        "navigation_error": False,
        "navigation_error_message": "",
    }
    result.update(overrides)
    return result


def _fake_row(i: int) -> dict:
    # 실제 apollo_list_collector.collect_apollo_first_list_query가 만드는
    # row와 동일한 형태(제품 13컬럼 + 내부 place_id/source_* 메타)를 흉내낸다.
    return {
        "업체명": f"업체{i}",
        "업종": "카페",
        "새로오픈여부": "",
        "방문자리뷰수": i,
        "블로그리뷰수": "",
        "총리뷰수": "",
        "주소": f"서울 강동구 천호동 {i}",
        "대표전화": "02-0000-0000",
        "플레이스 URL": f"https://pcmap.place.naver.com/place/{i}/home",
        "수집일": "2026-07-14",
        "홈페이지": "",
        "인스타": "",
        "블로그": "",
        "place_id": str(i),
        "source_city": "서울특별시",
        "source_district": "강동구",
        "source_subregion": "천호동",
        "source_layer": "legal_dong",
        "source_query": "서울특별시 강동구 천호동 카페",
    }


def _run_with_real_exporter(rows, final_count, output_path, *, stop_reason="target_reached", before_trim_count=None):
    app = _make_app()
    result = _base_result(
        stop_reason=stop_reason, final_count=final_count, rows=rows,
        before_trim_count=before_trim_count if before_trim_count is not None else final_count,
    )
    fake_orchestrator = _make_fake_orchestrator(result)
    return app._run_network_pipeline(
        [{"query": "q1"}], 30, 300, output_path,
        collector_factory=_fake_collector_factory, orchestrator=fake_orchestrator,
        excel_exporter=export_places_to_excel,
    )


def check_file_created_when_rows_present(reporter: ValidationReporter) -> None:
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "naver_place_network_db.xlsx")
        rows = [_fake_row(i) for i in range(3)]
        returned = _run_with_real_exporter(rows, 3, output_path)

        ok = Path(output_path).exists() and returned["exported"] is True and returned["export_path"] == output_path
        if ok:
            reporter.pass_("파일 생성: rows가 있으면 실제 .xlsx 파일이 생성되고 exported=True")
        else:
            reporter.fail(f"파일 생성 결과가 예상과 다름: exists={Path(output_path).exists()}, returned={returned}")


def check_final_row_count_matches_result(reporter: ValidationReporter) -> None:
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "out.xlsx")
        rows = [_fake_row(i) for i in range(5)]
        returned = _run_with_real_exporter(rows, 5, output_path)

        wb = openpyxl.load_workbook(output_path)
        data_row_count = wb["통합_결과"].max_row - 1  # 1행은 헤더
        ok = data_row_count == returned["final_count"] == 5
        if ok:
            reporter.pass_("최종 행 수: 통합_결과 시트의 데이터 행 수가 result['final_count']와 일치(5)")
        else:
            reporter.fail(f"최종 행 수 결과가 예상과 다름: data_row_count={data_row_count}, final_count={returned['final_count']}")


def check_header_matches_merged_columns(reporter: ValidationReporter) -> None:
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "out.xlsx")
        rows = [_fake_row(0)]
        _run_with_real_exporter(rows, 1, output_path)

        wb = openpyxl.load_workbook(output_path)
        headers = [cell.value for cell in wb["통합_결과"][1]]
        ok = headers == MERGED_COLUMNS and len(headers) == 14
        if ok:
            reporter.pass_("14컬럼: 통합_결과 헤더가 MERGED_COLUMNS와 정확히 일치하고 열 수 14개")
        else:
            reporter.fail(f"14컬럼 결과가 예상과 다름: headers={headers}")


def check_internal_meta_fields_not_in_excel(reporter: ValidationReporter) -> None:
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "out.xlsx")
        rows = [_fake_row(0)]
        _run_with_real_exporter(rows, 1, output_path)

        wb = openpyxl.load_workbook(output_path)
        headers = [cell.value for cell in wb["통합_결과"][1]]
        internal_fields = ("place_id", "source_city", "source_district", "source_subregion", "source_layer", "source_query")
        ok = all(field not in headers for field in internal_fields)
        if ok:
            reporter.pass_("내부 메타 미노출: place_id/source_city/source_district/source_subregion/source_layer/source_query 모두 헤더에 없음")
        else:
            reporter.fail(f"내부 메타 미노출 결과가 예상과 다름: headers={headers}")


def check_empty_original_sheets(reporter: ValidationReporter) -> None:
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "out.xlsx")
        rows = [_fake_row(0), _fake_row(1)]
        _run_with_real_exporter(rows, 2, output_path)

        wb = openpyxl.load_workbook(output_path)
        sheet_names_ok = set(wb.sheetnames) == {"통합_결과", "원본_모바일", "원본_PC"}
        mobile_empty = wb["원본_모바일"].max_row == 1  # 헤더만 있고 데이터 행 없음
        pc_empty = wb["원본_PC"].max_row == 1
        ok = sheet_names_ok and mobile_empty and pc_empty
        if ok:
            reporter.pass_("빈 원본 시트: 기존 3시트 구조 유지, 원본_모바일/원본_PC는 헤더만 있고 데이터 행 없음")
        else:
            reporter.fail(f"빈 원본 시트 결과가 예상과 다름: sheetnames={wb.sheetnames}, mobile_max_row={wb['원본_모바일'].max_row}, pc_max_row={wb['원본_PC'].max_row}")


def check_target_trim_row_count_consistency(reporter: ValidationReporter) -> None:
    """fake orchestrator가 target_count만큼 이미 trim된 rows(예: before_trim_count=5
    중 3개만 남김)를 반환하면, Excel 데이터 행도 정확히 그 trim된 수와 같아야 한다."""
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "out.xlsx")
        trimmed_rows = [_fake_row(i) for i in range(3)]
        returned = _run_with_real_exporter(
            trimmed_rows, final_count=3, output_path=output_path,
            stop_reason="target_reached", before_trim_count=5,
        )

        wb = openpyxl.load_workbook(output_path)
        data_row_count = wb["통합_결과"].max_row - 1
        ok = data_row_count == 3 == returned["final_count"] and returned["exported"] is True
        if ok:
            reporter.pass_("target trim 정합성: before_trim_count=5여도 실제 trim된 rows(3개)만큼 Excel에 저장됨")
        else:
            reporter.fail(f"target trim 정합성 결과가 예상과 다름: data_row_count={data_row_count}, returned={returned}")


def check_basic_mode_never_calls_home_enrichment_fn(reporter: ValidationReporter) -> None:
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "out.xlsx")
        app = _make_app()
        rows = [_fake_row(0)]
        result = _base_result(stop_reason="target_reached", final_count=1, rows=rows, before_trim_count=1)
        fake_orchestrator = _make_fake_orchestrator(result)

        home_calls: list = []

        def fake_home_enrichment_fn(rows, **kwargs):
            home_calls.append(rows)
            return {"rows": rows, "stop_reason": None, "security_blocked": False,
                    "home_success_count": 0, "failure_count": 0, "not_attempted_count": 0}

        app._run_network_pipeline(
            [{"query": "q1"}], 30, 300, output_path,
            collection_mode="basic",
            collector_factory=_fake_collector_factory, orchestrator=fake_orchestrator,
            excel_exporter=export_places_to_excel, home_enrichment_fn=fake_home_enrichment_fn,
        )
        if home_calls == []:
            reporter.pass_("기본 모드: home_enrichment_fn이 전혀 호출되지 않음")
        else:
            reporter.fail(f"기본 모드인데 home_enrichment_fn이 호출됨: {home_calls}")


def check_home_sns_mode_calls_home_enrichment_fn_once(reporter: ValidationReporter) -> None:
    """PAGE300-6A-FIX1: home_enrichment_fn은 더 이상 cookies를 인자로 받지
    않는다(Native Edge CDP persistent profile 재연결 방식으로 바뀌어 쿠키
    복사 자체가 불필요해졌으므로) - collector_factory는 기존 FakeNetworkCollector
    (capture_session_cookies 없음)를 그대로 재사용해도 무방하다."""
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "out.xlsx")
        app = _make_app()
        rows = [_fake_row(0)]
        result = _base_result(stop_reason="target_reached", final_count=1, rows=rows, before_trim_count=1)
        fake_orchestrator = _make_fake_orchestrator(result)

        home_calls: list = []
        enriched_row = dict(rows[0], 홈페이지="https://enriched.test")

        def fake_home_enrichment_fn(rows_arg, **kwargs):
            home_calls.append(list(rows_arg))
            return {"rows": [enriched_row], "stop_reason": None, "security_blocked": False,
                    "home_success_count": 1, "failure_count": 0, "not_attempted_count": 0}

        returned = app._run_network_pipeline(
            [{"query": "q1"}], 30, 300, output_path,
            collection_mode="home_sns",
            collector_factory=_fake_collector_factory, orchestrator=fake_orchestrator,
            excel_exporter=export_places_to_excel, home_enrichment_fn=fake_home_enrichment_fn,
        )

        wb = openpyxl.load_workbook(output_path)
        header = [cell.value for cell in wb["통합_결과"][1]]
        homepage_col = header.index("홈페이지") + 1
        saved_homepage = wb["통합_결과"].cell(row=2, column=homepage_col).value

        ok = (
            len(home_calls) == 1
            and returned["home_success_count"] == 1
            and saved_homepage == "https://enriched.test"
        )
        if ok:
            reporter.pass_("홈페이지·SNS 모드: home_enrichment_fn이 정확히 1회 호출되고 Excel에 보강된 rows가 반영됨")
        else:
            reporter.fail(f"홈페이지·SNS 모드 결과가 예상과 다름: home_calls={home_calls}, returned={returned}, saved_homepage={saved_homepage}")


def check_home_sns_diagnostics_saved_and_failure_does_not_block_export(reporter: ValidationReporter) -> None:
    """PAGE300-6G-R1: home_result에 diagnostics_report가 있으면
    ui_home_stage.save_json_artifact로 저장을 시도하고(UI-4: Home stage가
    src.ui_home_stage로 분리되며 저장 호출 지점도 함께 이동함), 저장
    자체가 실패해도(디스크 등 이유) Excel 저장 흐름은 막히지 않는다.
    최종 실패 업체 로그 라인도 함께 남는다."""
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "out.xlsx")
        app = _make_app()
        logs: list = []
        app.log = lambda message: logs.append(message)
        rows = [_fake_row(0)]
        result = _base_result(stop_reason="target_reached", final_count=1, rows=rows, before_trim_count=1)
        fake_orchestrator = _make_fake_orchestrator(result)

        enriched_row = dict(rows[0], 홈페이지="https://enriched.test")

        def fake_home_enrichment_fn(rows_arg, **kwargs):
            return {
                "rows": [enriched_row], "stop_reason": None, "security_blocked": False,
                "home_success_count": 0, "failure_count": 1, "not_attempted_count": 0,
                "first_pass": {"attempted": 1, "success": 0, "failed": 1},
                "retry_pass": {"attempted": 1, "success": 0, "failed": 1, "skipped": 0},
                "final_failures": [
                    {
                        "place_id": "0", "name": "업체0", "status": "timeout",
                        "http_status": None, "attempt": 2, "elapsed_ms": 15000,
                    }
                ],
                "diagnostics_report": {"run_id": "test-run", "target_count": 1},
            }

        save_calls: list = []

        def fake_save_json_artifact(run_dir, name, data):
            save_calls.append((str(run_dir), name, data))
            return DiagnosticArtifact(name=name, path=None, success=False, error_message="disk full")

        original_save = ui_home_stage.save_json_artifact
        ui_home_stage.save_json_artifact = fake_save_json_artifact
        try:
            returned = app._run_network_pipeline(
                [{"query": "q1"}], 30, 300, output_path,
                collection_mode="home_sns",
                collector_factory=_fake_collector_factory, orchestrator=fake_orchestrator,
                excel_exporter=export_places_to_excel, home_enrichment_fn=fake_home_enrichment_fn,
            )
        finally:
            ui_home_stage.save_json_artifact = original_save

        exported = Path(output_path).exists() and returned["exported"] is True
        diagnostics_attempted = len(save_calls) == 1 and save_calls[0][2].get("run_id") == "test-run"
        failure_logged = any("실패 1/1" in msg and "업체0" in msg and "timeout" in msg for msg in logs)
        save_failure_logged = any("진단 저장 실패" in msg for msg in logs)
        ok = exported and diagnostics_attempted and failure_logged and save_failure_logged
        if ok:
            reporter.pass_(
                "홈페이지·SNS 진단: diagnostics_report 저장 시도 + 최종 실패 로그 + "
                "저장 실패해도 Excel 저장은 그대로 성공"
            )
        else:
            reporter.fail(
                f"진단 저장/실패 로그 결과가 예상과 다름: exported={exported}, "
                f"diagnostics_attempted={diagnostics_attempted}, failure_logged={failure_logged}, "
                f"save_failure_logged={save_failure_logged}, logs={logs}"
            )


def check_zero_rows_no_file_created(reporter: ValidationReporter) -> None:
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "out.xlsx")
        returned = _run_with_real_exporter([], 0, output_path, stop_reason="queue_exhausted")

        ok = (
            not Path(output_path).exists()
            and returned["exported"] is False
            and returned["export_error"] is False
        )
        if ok:
            reporter.pass_("0건: _run_network_pipeline 정책에 따라 exporter가 호출되지 않고 파일이 생성되지 않음")
        else:
            reporter.fail(f"0건 결과가 예상과 다름: exists={Path(output_path).exists()}, returned={returned}")


def main() -> int:
    reporter = ValidationReporter()

    check_file_created_when_rows_present(reporter)
    check_final_row_count_matches_result(reporter)
    check_header_matches_merged_columns(reporter)
    check_internal_meta_fields_not_in_excel(reporter)
    check_empty_original_sheets(reporter)
    check_target_trim_row_count_consistency(reporter)
    check_basic_mode_never_calls_home_enrichment_fn(reporter)
    check_home_sns_mode_calls_home_enrichment_fn_once(reporter)
    check_home_sns_diagnostics_saved_and_failure_does_not_block_export(reporter)
    check_zero_rows_no_file_created(reporter)

    reporter.summary()
    return 1 if reporter.fail_count else 0


def test_standalone_suite():
    assert main() == 0


if __name__ == "__main__":
    sys.exit(main())
