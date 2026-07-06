from pathlib import Path
import sys
import tempfile

from openpyxl import load_workbook


# Stage 3C: src/pc/export_adapter.py(collect_pc_full row -> 통합_결과 직결) 검증 스크립트.
ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.exporter import MERGED_COLUMNS
from src.pc.export_adapter import export_full_collection


class ValidationReporter:
    def __init__(self):
        self.pass_count = 0
        self.fail_count = 0
        self.warn_count = 0

    def pass_(self, message: str) -> None:
        self.pass_count += 1
        print(f"[PASS] {message}")

    def fail(self, message: str) -> None:
        self.fail_count += 1
        print(f"[FAIL] {message}")

    def summary(self) -> None:
        final = "FAIL" if self.fail_count else "PASS"
        print("====================")
        print("검증 요약")
        print(f"PASS: {self.pass_count}")
        print(f"FAIL: {self.fail_count}")
        print(f"WARN: {self.warn_count}")
        print(f"FINAL: {final}")
        print("====================")


def _engine_rows():
    return [
        {
            "업체명": "오베르캄프 본점",
            "업종": "베이커리",
            "새로오픈여부": "",
            "리뷰수": "2469",
            "주소": "서울 강동구 성내로14길 48 1층",
            "대표전화": "0507-1387-4967",
            "플레이스 URL": "https://pcmap.place.naver.com/restaurant/1171815551/home",
            "place_id": "1171815551",
            "수집일": "2026-07-06",
            "홈페이지": "",
            "인스타": "https://www.instagram.com/oberkampf.kr",
            "블로그": "",
        }
    ]


def _read_sheet(path, sheet_name):
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name]
    headers = [cell.value for cell in worksheet[1]]
    rows = [dict(zip(headers, values)) for values in worksheet.iter_rows(min_row=2, values_only=True)]
    return headers, rows, workbook.sheetnames


def check_adapter_exports_merged(reporter: ValidationReporter) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "adapter_test.xlsx"
        returned = export_full_collection(_engine_rows(), str(path))

        headers, rows, sheet_names = _read_sheet(path, "통합_결과")
        row = rows[0] if rows else {}
        checks = {
            "저장 경로 반환 + 파일 존재": returned == str(path) and Path(path).exists(),
            "3시트 구조 유지": sheet_names == ["통합_결과", "원본_모바일", "원본_PC"],
            "통합_결과 헤더 = exporter.MERGED_COLUMNS": headers == MERGED_COLUMNS,
            "place_id 미노출": "place_id" not in headers,
            "대표전화 직결": row.get("대표전화") == "0507-1387-4967",
            "플레이스 URL 직결": row.get("플레이스 URL") == "https://pcmap.place.naver.com/restaurant/1171815551/home",
            "인스타 직결": row.get("인스타") == "https://www.instagram.com/oberkampf.kr",
        }
        if all(checks.values()):
            reporter.pass_("어댑터: collect_pc_full row를 parse_places 없이 통합_결과로 직결(11컬럼, place_id 미노출)")
        else:
            failed = [name for name, ok in checks.items() if not ok]
            reporter.fail(f"어댑터 검증 실패 항목: {failed} (headers={headers}, row={row})")


def check_adapter_empty_rows(reporter: ValidationReporter) -> None:
    # 빈 결과도 예외 없이 저장되어야 한다(부분 보존/빈 결과 흐름 호환).
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "adapter_empty.xlsx"
        try:
            returned = export_full_collection([], str(path))
        except Exception as exc:
            reporter.fail(f"빈 rows 저장에서 예외 발생: {type(exc).__name__}: {exc}")
            return
        headers, rows, _ = _read_sheet(path, "통합_결과")
        if Path(returned).exists() and headers == MERGED_COLUMNS and rows == []:
            reporter.pass_("빈 rows도 헤더만 있는 통합_결과로 정상 저장(예외 없음)")
        else:
            reporter.fail(f"빈 rows 저장 결과가 예상과 다름: headers={headers}, rows={rows}")


def main() -> int:
    reporter = ValidationReporter()
    check_adapter_exports_merged(reporter)
    check_adapter_empty_rows(reporter)
    reporter.summary()
    return 1 if reporter.fail_count else 0


if __name__ == "__main__":
    sys.exit(main())
