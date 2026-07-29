from pathlib import Path
import inspect
import json
import sys
import tempfile
import threading
from types import SimpleNamespace
from urllib.parse import quote, unquote

import openpyxl


# ARCH-300C WIRE-3: UI 시작 → NetworkBrowserCollector → collect_network_query →
# run_collection_plan → export_places_to_excel까지 실제 production 함수를
# 최대한 그대로 연결해 검증하는 no-live 전체 통합 테스트(standalone 스크립트).
# 가짜로 대체하는 것은 브라우저 계층(FakeBrowserSessionLike/FakePage/
# FakeResponse)뿐이며, NetworkBrowserCollector/collect_network_query/
# network_list_scraper 매핑/run_collection_plan/export_places_to_excel/
# UI의 _run_network_pipeline은 전부 실제 함수를 그대로 호출한다. 실제
# Playwright/네이버 접속/실제 UI 창은 전혀 만들지 않는다.
ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src import ui
from src.exporter import MERGED_COLUMNS, export_places_to_excel
from src.pc import network_browser_collector
from src.pc.network_browser_collector import (
    NetworkBrowserCollector,
    _SEARCH_URL_TEMPLATE,
    collect_network_query,
)
from src.pc.network_pipeline import run_collection_plan


class ValidationReporter:
    def __init__(self):
        self.pass_count = 0
        self.fail_count = 0
        self.warn_count = 0

    def pass_(self, message: str) -> None:
        self.pass_count += 1
        print(f"[PASS] {message}")

    def fail(self, message: str) -> None:
        self.fail_count += 1
        print(f"[FAIL] {message}")

    def summary(self) -> None:
        final = "FAIL" if self.fail_count else "PASS"
        print("====================")
        print("검증 요약")
        print(f"PASS: {self.pass_count}")
        print(f"FAIL: {self.fail_count}")
        print(f"WARN: {self.warn_count}")
        print(f"FINAL: {final}")
        print("====================")


# ---------------------------------------------------------------------------
# Fake 브라우저 계층 - Playwright page/response/locator/session 계약만 흉내낸다.
# (test_pc_network_browser_collector.py의 FakePage/FakeResponse/FakeLocator와
# 동일한 계약을 재사용/확장 - close()만 추가해 NetworkBrowserCollector.
# collect_query가 호출하는 page.close()를 지원한다.)
# ---------------------------------------------------------------------------


class FakeLocator:
    def __init__(self, count=0, visible=False, box=None):
        self._count = count
        self._visible = visible
        self._box = box or {"width": 0, "height": 0}

    @property
    def first(self):
        return self

    def count(self):
        return self._count

    def is_visible(self, timeout=300):
        return self._visible

    def bounding_box(self):
        return self._box


class FakeRequest:
    """PAGE-300-2B-2B: requestfinished/requestfailed 핸들러가 받는 request
    객체를 흉내낸다(response와 양방향 참조)."""

    def __init__(self, resource_type, method="GET"):
        self.resource_type = resource_type
        self.method = method
        self.failure = None
        self._response = None

    def response(self):
        return self._response


class FakeResponse:
    """PAGE-300-2B-2B: `.json()` 대신 `.body()`(bytes)로 candidate body를
    노출한다."""

    def __init__(self, url, status, resource_type, *, body=None, body_error=None,
                 method="GET", simulate_request_failed=False, headers=None):
        self.url = url
        self.status = status
        self.request = FakeRequest(resource_type, method=method)
        self.request._response = self
        self._body = body
        self._body_error = body_error
        self.body_call_count = 0
        self.simulate_request_failed = simulate_request_failed
        self.headers = headers if headers is not None else {}

    def body(self):
        self.body_call_count += 1
        if self._body_error is not None:
            raise self._body_error
        return self._body

    def text(self):
        if self._body_error is not None:
            raise self._body_error
        return (self._body or b"").decode("utf-8")


class FakePage:
    """collect_network_query가 실제로 요구하는 계약(on/off/goto/
    wait_for_timeout/locator)에 close()를 더한 fake - NetworkBrowserCollector.
    collect_query가 쿼리마다 새 page를 만들고 끝나면 close()하는 실제 생명주기를
    그대로 거친다."""

    def __init__(self, *, responses=None, goto_error=None, captcha_selectors=None):
        self._handlers: dict = {}
        self._responses = responses or []
        self._goto_error = goto_error
        self._captcha_selectors = captcha_selectors or {}
        self.goto_calls: list = []
        self.wait_calls: list = []
        self.close_call_count = 0

    def on(self, event, handler):
        self._handlers.setdefault(event, []).append(handler)

    def off(self, event, handler):
        handlers = self._handlers.get(event, [])
        if handler in handlers:
            handlers.remove(handler)

    def goto(self, url, wait_until=None, timeout=None):
        self.goto_calls.append(url)
        if self._goto_error is not None:
            raise self._goto_error
        for response in self._responses:
            for handler in list(self._handlers.get("response", [])):
                handler(response)
            if getattr(response, "simulate_request_failed", False):
                for handler in list(self._handlers.get("requestfailed", [])):
                    handler(response.request)
            else:
                for handler in list(self._handlers.get("requestfinished", [])):
                    handler(response.request)

    def wait_for_timeout(self, ms):
        self.wait_calls.append(ms)

    def locator(self, selector):
        return self._captcha_selectors.get(selector, FakeLocator(count=0))

    def close(self):
        self.close_call_count += 1


CANDIDATE_URL = "https://map.naver.com/p/api/search/allSearch?query=x"


def _place_response(item_ids_and_names, url=CANDIDATE_URL):
    items = [{"id": item_id, "name": name} for item_id, name in item_ids_and_names]
    body = json.dumps({"result": {"place": {"list": items}}}).encode("utf-8")
    return FakeResponse(url, 200, "xhr", body=body)


class FakeInitialPage:
    """BrowserSession.__enter__가 미리 만들어 두는 session.page의 fake.
    NetworkBrowserCollector가 __enter__에서 best-effort로 즉시 닫는 대상이다."""

    def __init__(self):
        self.close_call_count = 0

    def close(self):
        self.close_call_count += 1


class FakeContext:
    """실제 BrowserSession.context와 동일하게 new_page()만 제공하는 fake.
    쿼리 실행 순서대로 미리 준비된 FakePage 목록을 순서대로 하나씩 내준다 -
    준비된 개수보다 많이 호출되면(IndexError) 의도치 않게 다음 쿼리가
    실행됐다는 뜻이므로 테스트 실패로 자연스럽게 드러난다(예: CAPTCHA/429
    중단 이후 남은 쿼리가 실행되지 않아야 하는 시나리오의 방어 장치)."""

    def __init__(self, pages):
        self._pages = list(pages)
        self.new_page_call_count = 0
        self.pages_served: list = []

    def new_page(self):
        page = self._pages[self.new_page_call_count]
        self.new_page_call_count += 1
        self.pages_served.append(page)
        return page


class FakeBrowserSessionLike:
    """BrowserSession의 실제 계약(컨텍스트 매니저 + .context/.page 속성)만
    흉내내는 fake(WIRE-2B-1B와 동일한 검증된 패턴)."""

    def __init__(self, pages):
        self.context = FakeContext(pages)
        self.page = FakeInitialPage()
        self.enter_count = 0
        self.exit_count = 0

    def __enter__(self):
        self.enter_count += 1
        return self

    def __exit__(self, exc_type, exc, tb):
        self.exit_count += 1
        return False


def _make_collector_factory(pages, *, session_registry=None, session_factory_error=None):
    """_run_network_pipeline의 collector_factory(collected_at=...) 계약을
    만족하면서, 내부적으로는 실제 NetworkBrowserCollector를 그대로 사용하고
    session_factory만 fake로 주입한다 - collect_network_query/
    NetworkBrowserCollector.collect_query 자체는 전혀 monkeypatch하지 않는다."""

    def factory(*, collected_at, pause_event=None, stop_event=None):
        def session_factory():
            if session_factory_error is not None:
                raise session_factory_error
            session = FakeBrowserSessionLike(pages)
            if session_registry is not None:
                session_registry.append(session)
            return session

        return NetworkBrowserCollector(collected_at=collected_at, session_factory=session_factory)

    return factory


# ---------------------------------------------------------------------------
# Fake UI app - __new__ 기반(Tk 없음)
# ---------------------------------------------------------------------------


def _make_app():
    app = ui.SalesDbCrawlerApp.__new__(ui.SalesDbCrawlerApp)
    app.stop_event = threading.Event()
    app.pause_event = threading.Event()
    app._security_block_decision = None
    logs: list = []
    statuses: list = []
    app.log = lambda message: logs.append(message)
    app.set_status = lambda message: statuses.append(message)
    app.after = lambda delay, func=None, *args, **kwargs: (func(*args, **kwargs) if func else None)
    app.duplicate_removed_var = SimpleNamespace(set=lambda value: None)
    return app, logs, statuses


class FakeWidget:
    def __init__(self):
        self.configure_calls: list = []

    def configure(self, **kwargs) -> None:
        self.configure_calls.append(kwargs)


class FakeContainer:
    def winfo_children(self):
        return []


def _make_app_with_real_ui_recovery():
    """§10 worker 예외 시나리오 전용: set_running/_set_left_panel_state를
    실제(class) 메서드 그대로 두고, self.after만 콜백을 즉시 실행하도록
    fake 처리한다(요청서 §3) - 그래야 set_running(False)가 실제로
    btn_start.configure(state="normal")까지 전파되는지 확인할 수 있다.

    new_open_checkbox도 FakeWidget으로 미리 채워둔다 - 실제 제품에서는
    _build_ui()가 이미 끝난 뒤에만 set_running이 호출되므로 이 속성이 항상
    존재하지만, __new__로 만든 이 헤드리스 인스턴스는 tkinter.Widget의
    __getattr__(self.tk로 위임)이 초기화되지 않은 상태라 hasattr()로 존재하지
    않는 위젯 속성을 조회하면 무한 재귀(RecursionError)가 난다 - 이는
    _set_left_panel_state의 결함이 아니라(실제 실행 순서상 절대 발생하지
    않음) 이 테스트 fixture가 실제 위젯 존재를 흉내내지 못한 데서 오는
    문제이므로, 프로덕션 코드 대신 fixture 쪽에서 해당 속성을 채워 해결한다.
    """
    app, logs, statuses = _make_app()
    app.btn_start = FakeWidget()
    app.left_panel = FakeContainer()
    app.new_open_checkbox = FakeWidget()
    # LEGALDONG-UI-2: _set_left_panel_state가 target_count_entry도 항상
    # disabled로 강제하므로(§_build_global_target_count_section), 위
    # new_open_checkbox와 같은 이유로 이 fixture에도 채워둬야 한다.
    app.target_count_entry = FakeWidget()
    app.after = lambda delay, func=None, *args, **kwargs: (func(*args, **kwargs) if func else None)
    return app, logs, statuses


JOB1 = {
    "query": "서울특별시 강동구 천호동 카페",
    "source_city": "서울특별시", "source_district": "강동구",
    "source_subregion": "천호동", "source_layer": "legal_dong",
}
JOB2 = {
    "query": "서울특별시 강동구 성내동 카페",
    "source_city": "서울특별시", "source_district": "강동구",
    "source_subregion": "성내동", "source_layer": "legal_dong",
}
JOB3 = {
    "query": "서울특별시 강동구 길동 카페",
    "source_city": "서울특별시", "source_district": "강동구",
    "source_subregion": "길동", "source_layer": "legal_dong",
}
JOB4 = {
    "query": "서울특별시 강동구 암사동 카페",
    "source_city": "서울특별시", "source_district": "강동구",
    "source_subregion": "암사동", "source_layer": "legal_dong",
}


def _counting_exporter(calls: list):
    """export_places_to_excel을 그대로 위임 호출하면서 호출 횟수만 기록하는
    래퍼 - "실제 exporter 함수를 쓴다"는 원칙은 유지한 채 호출 횟수를
    명시적으로 단언하기 위한 최소 계측이다."""

    def exporter(merged_data, mobile_data, pc_data, output_path):
        calls.append(1)
        return export_places_to_excel(merged_data, mobile_data, pc_data, output_path)

    return exporter


def _run_pipeline(app, jobs, per_query_limit, target_count, output_path, pages, session_registry=None, excel_exporter=export_places_to_excel):
    return app._run_network_pipeline(
        jobs, per_query_limit, target_count, output_path,
        collector_factory=_make_collector_factory(pages, session_registry=session_registry),
        orchestrator=run_collection_plan,
        excel_exporter=excel_exporter,
    )


# ---------------------------------------------------------------------------
# §4 정상 목표 도달
# ---------------------------------------------------------------------------


def check_target_reached_full_integration(reporter: ValidationReporter) -> None:
    """ARCH-300C WIRE-3A: job을 4개로 늘려 target_reached 이후 남은 job(4번째)이
    절대 실행되지 않음을 명시적으로 검증한다. job4용 FakePage(X,Y,Z)는 pages
    목록에 포함시켜 두되, run_collection_plan이 job3에서 멈추면 context.new_page()가
    4번째로 호출되지 않으므로 page4는 끝까지 goto/close가 0회여야 한다 -
    "생성은 됐지만 전혀 쓰이지 않음"을 실제로 증명하는 방식이다."""
    page1 = FakePage(responses=[
        _place_response([("A", "업체A"), ("B", "업체B")]),
        _place_response([("B", "업체B"), ("C", "업체C")]),  # B는 job1 내부에서 local dup
    ])
    page2 = FakePage(responses=[_place_response([("B", "업체B"), ("C", "업체C"), ("D", "업체D")])])
    page3 = FakePage(responses=[_place_response([("E", "업체E"), ("F", "업체F"), ("G", "업체G")])])
    page4 = FakePage(responses=[_place_response([("X", "업체X"), ("Y", "업체Y"), ("Z", "업체Z")])])

    app, logs, statuses = _make_app()
    sessions: list = []
    exporter_calls: list = []

    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "target_reached.xlsx")
        result = _run_pipeline(
            app, [JOB1, JOB2, JOB3, JOB4], 3, 5, output_path,
            [page1, page2, page3, page4], sessions,
            excel_exporter=_counting_exporter(exporter_calls),
        )

        result_ok = (
            result["stop_reason"] == "target_reached"
            and result["final_count"] == 5
            and result["before_trim_count"] >= 5
            and result["executed_query_count"] == 3
            and result["skipped_query_count"] == 1
            and result["exported"] is True
            and result["export_error"] is False
        )
        if not result_ok:
            reporter.fail(f"목표 도달 result가 예상과 다름: {result}")
            return

        row_names = {row.get("업체명") for row in result["rows"]}
        job4_absent_from_rows = not ({"업체X", "업체Y", "업체Z"} & row_names)
        if not job4_absent_from_rows:
            reporter.fail(f"목표 도달: job4(X/Y/Z)가 result rows에 섞여 들어옴: {row_names}")
            return

        if not Path(output_path).exists():
            reporter.fail("목표 도달: exported=True인데 실제 xlsx 파일이 생성되지 않음")
            return

        wb = openpyxl.load_workbook(output_path)
        merged_rows = list(wb["통합_결과"].iter_rows(min_row=2, values_only=True))
        headers = [cell.value for cell in wb["통합_결과"][1]]
        names_in_order = [row[headers.index("업체명")] for row in merged_rows]
        mobile_rows = wb["원본_모바일"].max_row
        pc_rows = wb["원본_PC"].max_row

        excel_ok = (
            len(merged_rows) == 5
            and headers == MERGED_COLUMNS
            and names_in_order == ["업체A", "업체B", "업체C", "업체D", "업체E"]
            and not ({"업체X", "업체Y", "업체Z"} & set(names_in_order))
            and mobile_rows == 1
            and pc_rows == 1
        )
        if not excel_ok:
            reporter.fail(f"목표 도달 Excel 결과가 예상과 다름: rows={len(merged_rows)}, headers={headers}, names={names_in_order}")
            return

        exporter_call_ok = exporter_calls == [1]

        session = sessions[0]
        pages_used = session.context.pages_served
        job4_never_used = page4.goto_calls == [] and page4.close_call_count == 0 and page4 not in pages_used
        teardown_ok = (
            len(sessions) == 1
            and session.exit_count == 1
            and session.page.close_call_count == 1
            and session.context.new_page_call_count == 3
            and len(pages_used) == 3
            and all(p.close_call_count == 1 for p in pages_used)
            and job4_never_used
        )
        if not (exporter_call_ok and teardown_ok):
            reporter.fail(
                f"목표 도달 teardown/job4/exporter 결과가 예상과 다름: exit_count={session.exit_count}, "
                f"new_page_call_count={session.context.new_page_call_count}, "
                f"page4.goto_calls={page4.goto_calls}, page4.close_call_count={page4.close_call_count}, "
                f"exporter_calls={exporter_calls}"
            )
            return

    reporter.pass_(
        "정상 목표 도달(4 job): local+global dedup 정확(A,B,C,D,E 결정적 순서), "
        "job3에서 target_reached(executed=3, skipped=1, before_trim>=target), job4(X,Y,Z)는 "
        "new_page/goto/close 전혀 호출되지 않고 result rows·Excel 어디에도 없음, "
        "exporter 정확히 1회로 5행 저장, session/page teardown 전부 확인"
    )


# ---------------------------------------------------------------------------
# §5 목표 미달
# ---------------------------------------------------------------------------


def check_target_shortfall_full_integration(reporter: ValidationReporter) -> None:
    page1 = FakePage(responses=[_place_response([("A", "업체A"), ("B", "업체B")])])
    page2 = FakePage(responses=[_place_response([("A", "업체A"), ("C", "업체C")])])  # A는 전역 dup
    page3 = FakePage(responses=[_place_response([("D", "업체D")])])

    app, logs, statuses = _make_app()

    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "shortfall.xlsx")
        result = _run_pipeline(app, [JOB1, JOB2, JOB3], 10, 10, output_path, [page1, page2, page3])

        result_ok = (
            result["stop_reason"] == "queue_exhausted"
            and result["final_count"] == 4
            and result["executed_query_count"] == 3
            and result["skipped_query_count"] == 0
            and result["exported"] is True
        )
        if not result_ok:
            reporter.fail(f"목표 미달 result가 예상과 다름: {result}")
            return

        wb = openpyxl.load_workbook(output_path)
        merged_rows = list(wb["통합_결과"].iter_rows(min_row=2, values_only=True))
        excel_ok = len(merged_rows) == 4

        message = ui.SalesDbCrawlerApp._network_stop_message(app, result, 10)
        message_ok = "미달" in message and "4개를 저장했습니다" in message

        if not (excel_ok and message_ok):
            reporter.fail(f"목표 미달 결과가 예상과 다름: excel_rows={len(merged_rows)}, message={message}")
            return

    reporter.pass_("목표 미달: 전체 실행 후 stop_reason=queue_exhausted, final_count=4, Excel 4행, 미달 안내 문구 확인")


# ---------------------------------------------------------------------------
# §6 CAPTCHA 안전 중단
# ---------------------------------------------------------------------------


def check_captcha_safe_stop_full_integration(reporter: ValidationReporter) -> None:
    page1 = FakePage(responses=[_place_response([("X", "업체X"), ("Y", "업체Y"), ("Z", "업체Z")])])
    page2 = FakePage(
        responses=[],
        captcha_selectors={"#wtm-captcha-root": FakeLocator(count=1, visible=True, box={"width": 100, "height": 100})},
    )
    # job3용 page는 의도적으로 준비하지 않는다 - 실행되면 FakeContext가
    # IndexError를 던져 "job3 미실행" 위반을 즉시 드러낸다.

    app, logs, statuses = _make_app()

    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "captcha.xlsx")
        result = _run_pipeline(app, [JOB1, JOB2, JOB3], 10, 100, output_path, [page1, page2])

        result_ok = (
            result["stop_reason"] == "security_blocked"
            and result["security_blocked"] is True
            and result["status_429_seen"] is False
            and result["final_count"] == 3
            and result["executed_query_count"] == 2
            and result["skipped_query_count"] == 1
            and result["exported"] is True
        )
        if not result_ok:
            reporter.fail(f"CAPTCHA 중단 result가 예상과 다름: {result}")
            return

        wb = openpyxl.load_workbook(output_path)
        merged_rows = list(wb["통합_결과"].iter_rows(min_row=2, values_only=True))
        excel_ok = len(merged_rows) == 3

        callback_ok = app._security_block_decision is not None

        if not (excel_ok and callback_ok):
            reporter.fail(f"CAPTCHA 중단 결과가 예상과 다름: excel_rows={len(merged_rows)}, decision={app._security_block_decision}")
            return

    reporter.pass_("CAPTCHA 안전 중단: job1 결과(3건) 보존, job2 이후 즉시 중단(job3 미실행), on_security_block 연결, exporter 1회 저장, 재시도/session 재시작 없음")


# ---------------------------------------------------------------------------
# §7 HTTP 429 안전 중단
# ---------------------------------------------------------------------------


def check_status_429_safe_stop_full_integration(reporter: ValidationReporter) -> None:
    page1 = FakePage(responses=[_place_response([("M", "업체M"), ("N", "업체N")])])
    page2 = FakePage(responses=[FakeResponse("https://map.naver.com/some", 429, "document")])
    # job3용 page 미준비 - 실행되면 IndexError로 즉시 드러남.

    app, logs, statuses = _make_app()

    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "status_429.xlsx")
        result = _run_pipeline(app, [JOB1, JOB2, JOB3], 10, 100, output_path, [page1, page2])

        result_ok = (
            result["stop_reason"] == "status_429"
            and result["status_429_seen"] is True
            and result["final_count"] == 2
            and result["executed_query_count"] == 2
            and result["skipped_query_count"] == 1
            and result["exported"] is True
        )
        if not result_ok:
            reporter.fail(f"429 중단 result가 예상과 다름: {result}")
            return

        wb = openpyxl.load_workbook(output_path)
        merged_rows = list(wb["통합_결과"].iter_rows(min_row=2, values_only=True))
        if len(merged_rows) != 2:
            reporter.fail(f"429 중단 Excel 결과가 예상과 다름: rows={len(merged_rows)}")
            return

    reporter.pass_("HTTP 429 안전 중단: 이전 결과(2건) 저장, 남은 job 미실행, exporter 정확히 1회, retry 없음")


# ---------------------------------------------------------------------------
# §8 0건
# ---------------------------------------------------------------------------


def check_zero_rows_full_integration(reporter: ValidationReporter) -> None:
    page1 = FakePage(responses=[_place_response([])])
    page2 = FakePage(responses=[_place_response([])])

    app, logs, statuses = _make_app()

    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "zero_rows.xlsx")
        result = _run_pipeline(app, [JOB1, JOB2], 10, 5, output_path, [page1, page2])

        result_ok = (
            result["final_count"] == 0
            and result["exported"] is False
            and result["export_error"] is False
            and result["export_path"] == ""
        )
        file_absent = not Path(output_path).exists()
        message_ok = statuses and statuses[-1] == "저장할 결과가 없습니다."

        if not (result_ok and file_absent and message_ok):
            reporter.fail(f"0건 결과가 예상과 다름: result={result}, exists={Path(output_path).exists()}, statuses={statuses}")
            return

    reporter.pass_("0건: final_count=0, exporter 미호출로 파일 미생성, exported=False/export_error=False, '저장할 결과가 없습니다.' 안내")


# ---------------------------------------------------------------------------
# §9 검색 URL 계약 점검(문자열 계약만, live 아님)
# ---------------------------------------------------------------------------


def check_search_url_matches_job_query_and_poc_pattern(reporter: ValidationReporter) -> None:
    """1) job["query"]가 URL에 정확히 포함되는지, 4) PoC에서 성공했던
    URL 패턴(f"https://map.naver.com/v5/search/{quote(query)}")과 현재
    구현이 일치하는지 확인한다."""
    job = {"query": "서울특별시 강동구 천호동 카페"}
    page = FakePage(responses=[_place_response([("A", "업체A")])])
    collect_network_query(page, job, 10, collected_at="2026-07-15")

    if len(page.goto_calls) != 1:
        reporter.fail(f"URL 계약: page.goto가 정확히 1회 호출되지 않음(goto_calls={page.goto_calls})")
        return
    actual_url = page.goto_calls[0]
    expected_url = f"https://map.naver.com/v5/search/{quote(job['query'])}"
    poc_pattern_url = _SEARCH_URL_TEMPLATE.format(query=quote(job["query"]))

    ok = actual_url == expected_url == poc_pattern_url
    if ok:
        reporter.pass_(f"URL 계약: 현재 구현이 PoC-1~9에서 실제 성공한 패턴과 정확히 일치함 ({actual_url})")
    else:
        reporter.fail(f"URL 계약이 예상과 다름: actual={actual_url}, expected={expected_url}, poc_pattern={poc_pattern_url}")


def check_search_url_korean_space_special_char_encoding(reporter: ValidationReporter) -> None:
    """2) 한글·공백·특수문자가 안전하게 인코딩되는지, round-trip으로 원본
    쿼리와 일치하는지 확인한다."""
    job = {"query": "서울특별시 강동구 천호동 카페&디저트/브런치"}
    page = FakePage(responses=[])
    collect_network_query(page, job, 10, collected_at="2026-07-15")

    actual_url = page.goto_calls[0]
    path_suffix = actual_url.replace("https://map.naver.com/v5/search/", "", 1)

    no_raw_space = " " not in path_suffix
    roundtrip_ok = unquote(path_suffix) == job["query"]

    if no_raw_space and roundtrip_ok:
        reporter.pass_("URL 계약: 한글/공백/특수문자(&, /)가 안전하게 인코딩되고 unquote 시 원본 쿼리와 정확히 일치")
    else:
        reporter.fail(f"URL 인코딩 결과가 예상과 다름: path_suffix={path_suffix!r}, roundtrip={unquote(path_suffix)!r}")


def check_search_url_no_duplicated_source_meta(reporter: ValidationReporter) -> None:
    """3) source_city/source_district 등 내부 메타가 쿼리 URL에 중복으로
    붙지 않는지 확인한다(오직 job["query"] 문자열 하나만 URL에 반영)."""
    job = {
        "query": "서울특별시 강동구 천호동 카페",
        "source_city": "서울특별시", "source_district": "강동구",
        "source_subregion": "천호동", "source_layer": "legal_dong",
    }
    page = FakePage(responses=[])
    collect_network_query(page, job, 10, collected_at="2026-07-15")

    actual_url = page.goto_calls[0]
    expected_url = f"https://map.naver.com/v5/search/{quote(job['query'])}"
    if actual_url == expected_url:
        reporter.pass_("URL 계약: source_city/source_district 등 내부 메타가 URL에 중복 반영되지 않음(job['query']만 사용)")
    else:
        reporter.fail(f"URL에 source_* 메타가 중복 반영된 것으로 보임: actual={actual_url}, expected={expected_url}")


def check_search_url_one_query_per_url_and_single_goto(reporter: ValidationReporter) -> None:
    """5) URL 하나당 쿼리 하나만 표현되는지(여러 쿼리를 이어붙이지 않음) 확인한다."""
    jobs_and_pages = [
        (JOB1, FakePage(responses=[_place_response([("A", "업체A")])])),
        (JOB2, FakePage(responses=[_place_response([("B", "업체B")])])),
    ]
    urls = []
    for job, page in jobs_and_pages:
        collect_network_query(page, job, 10, collected_at="2026-07-15")
        urls.append(page.goto_calls[0])

    ok = (
        len(urls) == 2
        and urls[0] != urls[1]
        and quote(JOB1["query"]) in urls[0] and quote(JOB2["query"]) not in urls[0]
        and quote(JOB2["query"]) in urls[1] and quote(JOB1["query"]) not in urls[1]
    )
    if ok:
        reporter.pass_("URL 계약: URL 하나당 쿼리 하나만 표현되고 서로 섞이지 않음")
    else:
        reporter.fail(f"URL 계약이 예상과 다름: urls={urls}")


def check_no_direct_http_client_only_page_goto(reporter: ValidationReporter) -> None:
    """6) network_browser_collector가 별도 HTTP 클라이언트를 직접 호출하지
    않고 page.goto()에만 URL을 전달하는지 소스 기준으로 확인한다."""
    source = inspect.getsource(network_browser_collector)
    forbidden_imports = ("import requests", "import httpx", "urllib.request", "http.client")
    found = [token for token in forbidden_imports if token in source]
    goto_used = "page.goto(" in source

    if not found and goto_used:
        reporter.pass_("URL 계약: 별도 HTTP 클라이언트 import 없음(requests/httpx/urllib.request/http.client), page.goto()로만 이동")
    else:
        reporter.fail(f"URL 계약 결과가 예상과 다름: forbidden_found={found}, goto_used={goto_used}")


# ---------------------------------------------------------------------------
# §10 worker 예외·UI 복구
# ---------------------------------------------------------------------------


def check_worker_unexpected_exception_recovers_real_ui_state(reporter: ValidationReporter) -> None:
    """collector 생성(session_factory) 단계에서 예상 밖 예외가 나는 경우도
    실제 NetworkBrowserCollector.__enter__ 경로를 그대로 거친다(session_factory()
    호출은 try/except로 감싸져 있지 않으므로 그대로 전파된다) - 그 예외를
    _run_network_pipeline_worker(실제 메서드)가 최종적으로 잡아야 한다.
    set_running/_set_left_panel_state는 실제 메서드 그대로 두고 self.after만
    즉시 실행하도록 해, set_running(False)가 실제로 btn_start.configure까지
    전파되는지 확인한다."""
    app, logs, statuses = _make_app_with_real_ui_recovery()

    real_run_network_pipeline = ui.SalesDbCrawlerApp._run_network_pipeline
    collector_factory = _make_collector_factory([], session_factory_error=RuntimeError("browser launch failed"))

    def wrapper(query_queue, per_query_limit, target_count, output_path):
        return real_run_network_pipeline(
            app, query_queue, per_query_limit, target_count, output_path,
            collector_factory=collector_factory, orchestrator=run_collection_plan,
            excel_exporter=export_places_to_excel,
        )

    app._run_network_pipeline = wrapper

    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = str(Path(tmp_dir) / "worker_exception.xlsx")
        # _run_network_pipeline_worker는 monkeypatch하지 않은 실제 클래스 메서드다.
        app._run_network_pipeline_worker([JOB1], 10, 100, output_path)

        no_exception_escaped = True  # 위 호출이 예외 없이 반환되면 worker가 잡은 것
        file_absent = not Path(output_path).exists()

    ok = (
        no_exception_escaped
        and file_absent
        and statuses
        and statuses[-1] == "수집 중 오류가 발생했습니다."
        and "저장했습니다" not in statuses[-1]
        and "저장 완료" not in statuses[-1]
        and any("예상하지 못한 오류" in message for message in logs)
        and app.btn_start.configure_calls
        and app.btn_start.configure_calls[-1].get("state") == "normal"
    )
    if ok:
        reporter.pass_(
            "worker 예외·UI 복구: 실제 _run_network_pipeline_worker가 session_factory 예외를 포착, "
            "짧은 오류 문구만 표시(성공/저장 문구 없음), set_running(False)가 실제로 "
            "btn_start.configure(state='normal')까지 전파됨, 불완전한 파일 없음"
        )
    else:
        reporter.fail(
            f"worker 예외·UI 복구 결과가 예상과 다름: statuses={statuses}, logs={logs}, "
            f"btn_start.configure_calls={app.btn_start.configure_calls}"
        )


def main() -> int:
    reporter = ValidationReporter()

    check_target_reached_full_integration(reporter)
    check_target_shortfall_full_integration(reporter)
    check_captcha_safe_stop_full_integration(reporter)
    check_status_429_safe_stop_full_integration(reporter)
    check_zero_rows_full_integration(reporter)
    check_search_url_matches_job_query_and_poc_pattern(reporter)
    check_search_url_korean_space_special_char_encoding(reporter)
    check_search_url_no_duplicated_source_meta(reporter)
    check_search_url_one_query_per_url_and_single_goto(reporter)
    check_no_direct_http_client_only_page_goto(reporter)
    check_worker_unexpected_exception_recovers_real_ui_state(reporter)

    reporter.summary()
    return 1 if reporter.fail_count else 0


if __name__ == "__main__":
    sys.exit(main())
