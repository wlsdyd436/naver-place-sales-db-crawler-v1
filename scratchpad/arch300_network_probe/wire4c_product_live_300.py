# ARCH-300C WIRE-4C: 실제 Playwright + 네이버로 production 코드 경로(NetworkBrowserCollector
# -> collect_network_query -> run_collection_plan -> export_places_to_excel)를 검색
# 조합 최대 24개, 목표 300건으로 단발성 실행하는 live harness. WIRE-4B harness
# 구조를 재사용하되, Tier1(법정동)/Tier2(역·상권)/Tier3(세부업종) 3단 큐와
# tier별 진단 집계를 추가했다.
#
# 큐 구성 근거(§작업 전 확인에서 대조):
# - Tier1 법정동 9개, Tier2 역/상권 6개는 data/regions_kr_sample.json(현재 실제
#   지역 데이터, src/pc/region_data.py가 읽는 파일)의 강동구 legal_dongs/
#   landmarks 리스트 순서와 정확히 일치하며, 이는 scratchpad/arch300_network_probe/
#   poc7_target_300_probe.py의 LEGAL_DONGS/LANDMARKS 순서와도 동일하다(PoC-7
#   live 실행에서 CAPTCHA/429 없이 17개 쿼리 성공 검증된 순서).
# - Tier3 세부업종: work order 본문에 나열된 목록(천호동/길동/성내동 x
#   디저트카페/브런치카페, 6개)을 실제 poc7_target_300_probe.py의
#   SUBCATEGORY_DONGS=["천호동","성내동","길동"] / SUBCATEGORIES=["디저트카페",
#   "브런치카페","베이커리카페"]와 대조한 결과, "베이커리카페" 3개 job이
#   누락되어 있었다(PoC-7 실제 24개 큐 = Tier1 9 + Tier2 6 + Tier3 9, 이번
#   work order 본문 목록은 9가 아닌 6개만 나열됨). data/regions_kr_sample.json
#   에도 subcategory_keywords가 ["디저트카페","브런치카페","베이커리카페"] 3개로
#   정의되어 있어, 현재 지역 데이터와도 9개가 일치한다. 따라서 이 harness는
#   PoC-7이 실제로 성공한 24개(9+6+9) 전체 큐를 사용한다 - "빠진 세부업종 job이
#   있는지 확인하라"는 지시에 따라 발견한 차이를 반영해 큐를 완성했다(live
#   실행 전에 고정, 실행 중 수정 없음).
# - Tier1 순서도 work order 본문(가나다순 나열)과 실제 poc7_target_300_probe.py
#   LEGAL_DONGS 순서(천호동이 첫 항목)가 달라, "PoC-7 성공 순서를 기준으로"라는
#   §3 지시에 따라 poc7_target_300_probe.py에 기록된 실제 순서를 채택했다.
#   Tier2 순서는 work order 본문과 poc7_target_300_probe.py가 이미 동일하다.
# - job dict의 키 이름(source_city/source_district/source_subregion/source_layer)은
#   src/pc/network_browser_collector.py::collect_network_query가 실제로
#   job.get(...)하는 키와 정확히 일치시켰다(region_expander.py의 PoC 전용
#   빌더는 city/gu/dong 등 다른 키 이름을 쓰므로 그대로 재사용하면 메타가
#   비게 되어 사용하지 않았다 - 대신 검색어 문자열/순서만 대조 근거로 삼았다).
#
# 이 스크립트는 정확히 1회만 실행하도록 설계됐다 - 내부에 재시도/재검색/CAPTCHA
# 우회/context 재시작/병렬 실행 로직이 전혀 없다. fake session/page/response는
# 전혀 쓰지 않으며, 실제 production 함수를 그대로 호출한다. production 코드는
# 이 스크립트에서 절대 수정하지 않으며, settle_ms(5000, 기본값)도 변경하지 않는다.
from datetime import datetime
from pathlib import Path
from urllib.parse import quote
import json
import sys
import time

import openpyxl

ROOT_DIR = Path(__file__).resolve().parents[2]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.exporter import MERGED_COLUMNS, export_places_to_excel
from src.pc.browser_session import BrowserSession
from src.pc.config import DiagnosticConfig
from src.pc.network_browser_collector import NetworkBrowserCollector, _SEARCH_URL_TEMPLATE
from src.pc.network_list_scraper import classify_query_efficiency, dedup_rows as diagnostic_dedup_rows
from src.pc.network_pipeline import run_collection_plan

_COMMON_META = {"source_city": "서울특별시", "source_district": "강동구"}

# Tier1: 법정동(legal_dong) - data/regions_kr_sample.json / poc7_target_300_probe.py
# LEGAL_DONGS와 순서까지 정확히 동일.
_TIER1_LEGAL_DONGS = ["천호동", "성내동", "길동", "암사동", "명일동", "고덕동", "상일동", "둔촌동", "강일동"]
# Tier2: 역/상권(landmark) - 동일 출처, work order 본문과도 이미 순서 일치.
_TIER2_LANDMARKS = ["천호역", "강동역", "둔촌동역", "암사역", "고덕역", "명일역"]
# Tier3: 세부업종(vertical) - poc7_target_300_probe.py SUBCATEGORY_DONGS/SUBCATEGORIES
# 순서 그대로(동 우선 순회 후 업종 순회, work order 본문에서 누락됐던 "베이커리카페" 포함).
_TIER3_DONGS = ["천호동", "성내동", "길동"]
_TIER3_SUBCATEGORIES = ["디저트카페", "브런치카페", "베이커리카페"]

JOBS: list = []
for dong in _TIER1_LEGAL_DONGS:
    JOBS.append({
        "tier": "tier1", "source_layer": "legal_dong",
        "source_subregion": dong, "query": f"서울특별시 강동구 {dong} 카페", **_COMMON_META,
    })
for landmark in _TIER2_LANDMARKS:
    JOBS.append({
        "tier": "tier2", "source_layer": "landmark",
        "source_subregion": landmark, "query": f"서울특별시 강동구 {landmark} 카페", **_COMMON_META,
    })
for dong in _TIER3_DONGS:
    for subcategory in _TIER3_SUBCATEGORIES:
        JOBS.append({
            "tier": "tier3", "source_layer": "vertical",
            "source_subregion": dong, "query": f"서울특별시 강동구 {dong} {subcategory}", **_COMMON_META,
        })

assert len(JOBS) == 24, f"큐 구성 오류: 예상 24개, 실제 {len(JOBS)}개"

PER_QUERY_LIMIT = 20
TARGET_COUNT = 300

RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_XLSX_PATH = ROOT_DIR / "output" / f"wire4c_live_300_{RUN_TIMESTAMP}.xlsx"
RESULT_JSON_PATH = Path(__file__).resolve().parent / f"wire4c_live_300_result_{RUN_TIMESTAMP}.json"


def _safe_session_factory():
    # "기본 BrowserSession 경로"를 명시적으로 사용한다(고객용 안전 모드 -
    # visible=False/capture_artifacts=False/keep_open_on_error=False).
    return BrowserSession(DiagnosticConfig.safe_default())


def main() -> int:
    collected_at = datetime.now().strftime("%Y-%m-%d")

    timings: dict = {}
    per_query_log: list = []
    diagnostic_seen: set = set()
    cumulative_unique = 0

    def make_recording_collect_query(collector):
        """collector.collect_query를 그대로 호출하되(동작 변경 없음), 실행 순번/
        tier/진단값을 harness 보고용으로만 기록한다. global unique_added는
        진단 전용 seen set(diagnostic_seen)에만 적용하며 run_collection_plan에는
        전달하지 않는다 - 실제 오케스트레이션 dedup 로직과 완전히 분리된
        관찰용 계산이다."""

        def wrapper(job, per_query_limit):
            nonlocal cumulative_unique
            index = len(per_query_log) + 1
            search_url = _SEARCH_URL_TEMPLATE.format(query=quote(job["query"]))
            t0 = time.perf_counter()
            result = collector.collect_query(job, per_query_limit)
            t1 = time.perf_counter()

            rows = result.get("rows") or []
            newly_unique = diagnostic_dedup_rows(rows, diagnostic_seen)
            cumulative_unique += len(newly_unique)
            raw_item_count = result.get("raw_item_count") or 0
            efficiency = classify_query_efficiency(raw_item_count, len(newly_unique))

            per_query_log.append({
                "index": index,
                "tier": job.get("tier"),
                "source_layer": job.get("source_layer"),
                "query": job["query"],
                "search_url": search_url,
                "wall_seconds": t1 - t0,
                "candidate_response_count": result.get("candidate_response_count"),
                "raw_item_count": raw_item_count,
                "local_unique_count": result.get("local_unique_count"),
                "returned_rows_count": len(rows),
                "global_unique_added_diagnostic": len(newly_unique),
                "global_duplicate_count_diagnostic": len(rows) - len(newly_unique),
                "cumulative_unique_diagnostic": cumulative_unique,
                "efficiency_ratio_diagnostic": efficiency["efficiency_ratio"],
                "parse_error_count": result.get("parse_error_count"),
                "timeout": result.get("timeout"),
                "active_captcha_detected": result.get("active_captcha_detected"),
                "status_429_seen": result.get("status_429_seen"),
                "navigation_error": result.get("navigation_error"),
                "navigation_error_message": (result.get("navigation_error_message") or "")[:200],
            })
            print(f"[wire4c] query {index}/{len(JOBS)} [{job.get('tier')}] 완료: {job['query']!r} "
                  f"candidate={result.get('candidate_response_count')} raw={raw_item_count} "
                  f"local_unique={result.get('local_unique_count')} returned_rows={len(rows)} "
                  f"global_unique_added={len(newly_unique)} cumulative={cumulative_unique} "
                  f"wall={t1 - t0:.2f}s")
            return result

        return wrapper

    print(f"[wire4c] job 수={len(JOBS)}(tier1={len(_TIER1_LEGAL_DONGS)}, "
          f"tier2={len(_TIER2_LANDMARKS)}, tier3={len(_TIER3_DONGS) * len(_TIER3_SUBCATEGORIES)}), "
          f"per_query_limit={PER_QUERY_LIMIT}, target_count={TARGET_COUNT}")
    print("[wire4c] live 실행 시작(정확히 1회, 재시도 없음, settle_ms=5000 고정 유지)")

    t_start = time.perf_counter()
    export_result_path = None
    export_error_message = ""
    exported = False

    with NetworkBrowserCollector(collected_at=collected_at, session_factory=_safe_session_factory) as collector:
        t_session_ready = time.perf_counter()
        timings["session_ready_seconds"] = t_session_ready - t_start

        orchestrator_result = run_collection_plan(
            JOBS,
            per_query_limit=PER_QUERY_LIMIT,
            target_count=TARGET_COUNT,
            collected_at=collected_at,
            collect_query=make_recording_collect_query(collector),
        )
        t_orchestrator_done = time.perf_counter()
        timings["orchestrator_seconds"] = t_orchestrator_done - t_session_ready

    t_session_closed = time.perf_counter()
    timings["session_teardown_seconds"] = t_session_closed - t_orchestrator_done

    rows = orchestrator_result.get("rows") or []
    if rows:
        try:
            saved_path = export_places_to_excel(rows, [], [], str(OUTPUT_XLSX_PATH))
            exported = True
            export_result_path = str(saved_path or OUTPUT_XLSX_PATH)
        except Exception as exc:
            export_error_message = f"{type(exc).__name__}: {exc}"
            print(f"[wire4c] Excel 저장 실패: {export_error_message}")
        t_export_done = time.perf_counter()
        timings["export_seconds"] = t_export_done - t_session_closed
    else:
        timings["export_seconds"] = 0.0
        print("[wire4c] rows=0 - export_places_to_excel 호출하지 않음(파일 미생성)")

    excel_verification: dict = {}
    if exported and export_result_path:
        wb = openpyxl.load_workbook(export_result_path)
        merged_ws = wb["통합_결과"]
        headers = [cell.value for cell in merged_ws[1]]
        merged_data_rows = list(merged_ws.iter_rows(min_row=2, values_only=True))
        mobile_ws = wb["원본_모바일"]
        pc_ws = wb["원본_PC"]

        internal_fields = ("place_id", "source_city", "source_district", "source_subregion", "source_layer", "source_query")
        header_leak = [field for field in internal_fields if field in headers]

        def _count_filled(col_name: str):
            if col_name not in headers:
                return None
            idx = headers.index(col_name)
            return sum(1 for row in merged_data_rows if row[idx] not in (None, ""))

        excel_verification = {
            "headers": headers,
            "headers_match_merged_columns": headers == MERGED_COLUMNS,
            "data_row_count": len(merged_data_rows),
            "data_row_count_matches_final_count": len(merged_data_rows) == orchestrator_result.get("final_count"),
            "internal_field_leak": header_leak,
            "mobile_sheet_max_row": mobile_ws.max_row,
            "pc_sheet_max_row": pc_ws.max_row,
            "field_fill_counts": {col: _count_filled(col) for col in MERGED_COLUMNS},
        }

    t_end = time.perf_counter()
    timings["total_wall_seconds"] = t_end - t_start
    executed_count = len(per_query_log)
    timings["avg_query_wall_seconds"] = (
        sum(entry["wall_seconds"] for entry in per_query_log) / executed_count if executed_count else 0.0
    )
    final_count = orchestrator_result.get("final_count") or 0
    timings["rows_per_second"] = (final_count / timings["total_wall_seconds"]) if timings["total_wall_seconds"] else 0.0
    timings["seconds_per_final_row"] = (timings["total_wall_seconds"] / final_count) if final_count else None
    timings["estimated_settle_seconds"] = executed_count * 5.0  # settle_ms=5000 고정, 변경 없음

    total_raw_items = sum((entry["raw_item_count"] or 0) for entry in per_query_log)
    total_returned_rows = sum(entry["returned_rows_count"] for entry in per_query_log)
    before_trim_count = orchestrator_result.get("before_trim_count") or 0
    global_duplicates_removed = total_returned_rows - before_trim_count
    global_dedup_rate = (global_duplicates_removed / total_returned_rows) if total_returned_rows else 0.0
    diagnostic_matches_orchestrator = cumulative_unique == before_trim_count

    tier_summary: dict = {}
    for entry in per_query_log:
        tier = entry["tier"] or "unknown"
        bucket = tier_summary.setdefault(tier, {
            "executed_count": 0, "raw_total": 0, "local_unique_total": 0,
            "returned_rows_total": 0, "global_unique_added_total": 0,
        })
        bucket["executed_count"] += 1
        bucket["raw_total"] += entry["raw_item_count"] or 0
        bucket["local_unique_total"] += entry["local_unique_count"] or 0
        bucket["returned_rows_total"] += entry["returned_rows_count"]
        bucket["global_unique_added_total"] += entry["global_unique_added_diagnostic"]

    lowest_efficiency_query = None
    if per_query_log:
        lowest_efficiency_query = min(
            per_query_log, key=lambda entry: entry["efficiency_ratio_diagnostic"]
        )
        lowest_efficiency_query = {
            "index": lowest_efficiency_query["index"],
            "tier": lowest_efficiency_query["tier"],
            "query": lowest_efficiency_query["query"],
            "efficiency_ratio_diagnostic": lowest_efficiency_query["efficiency_ratio_diagnostic"],
            "global_unique_added_diagnostic": lowest_efficiency_query["global_unique_added_diagnostic"],
        }

    summary = {
        "total_job_count": len(JOBS),
        "per_query_limit": PER_QUERY_LIMIT,
        "target_count": TARGET_COUNT,
        "executed_query_count": orchestrator_result.get("executed_query_count"),
        "skipped_query_count": orchestrator_result.get("skipped_query_count"),
        "before_trim_count": orchestrator_result.get("before_trim_count"),
        "final_count": orchestrator_result.get("final_count"),
        "stop_reason": orchestrator_result.get("stop_reason"),
        "security_blocked": orchestrator_result.get("security_blocked"),
        "status_429_seen": orchestrator_result.get("status_429_seen"),
        "navigation_error": orchestrator_result.get("navigation_error"),
        "navigation_error_message": (orchestrator_result.get("navigation_error_message") or "")[:200],
        "parse_error_count_total": sum((entry["parse_error_count"] or 0) for entry in per_query_log),
        "exported": exported,
        "export_path": export_result_path or "",
        "export_error": bool(export_error_message),
        "export_error_message": export_error_message,
        "total_raw_items": total_raw_items,
        "total_local_unique": sum((entry["local_unique_count"] or 0) for entry in per_query_log),
        "total_returned_rows_after_local_dedup_and_cap": total_returned_rows,
        "global_duplicates_removed": global_duplicates_removed,
        "global_dedup_rate": global_dedup_rate,
        "diagnostic_cumulative_unique_matches_orchestrator_before_trim": diagnostic_matches_orchestrator,
        "tier_summary": tier_summary,
        "lowest_efficiency_query_diagnostic": lowest_efficiency_query,
        "timings": timings,
        "per_query_log": per_query_log,
        "excel_verification": excel_verification,
        "retries": 0,
    }

    print("[wire4c] ==== 결과 요약 ====")
    for key, value in summary.items():
        if key in ("timings", "excel_verification", "per_query_log", "tier_summary"):
            continue
        print(f"[wire4c] {key}={value}")
    print(f"[wire4c] tier_summary={tier_summary}")
    print(f"[wire4c] timings={timings}")
    print(f"[wire4c] excel_verification={excel_verification}")

    RESULT_JSON_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[wire4c] 결과 JSON 저장: {RESULT_JSON_PATH}")

    if exported:
        print(f"[wire4c] Excel 저장 경로: {export_result_path}")

    print("[wire4c] 실행 종료(재실행하지 않음)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
