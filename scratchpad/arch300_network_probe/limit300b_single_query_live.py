# ARCH-300C LIMIT-300-B: 검색 조합 1개("서울특별시 강동구 카페")에 대해
# per_query_limit=300/target_count=300을 실제 production 경로
# (NetworkBrowserCollector -> collect_network_query -> run_collection_plan
# -> export_places_to_excel)로 정확히 1회 실행해, 현재 코드가 스크롤/
# 페이지네이션 추가 없이 실제로 몇 건을 확보하는지 측정하는 live harness.
# perf1c_adaptive_live_300.py(실행 마커/설정 검증/exporter counting
# wrapper/export 전 중복 검증/Excel 검증 구조)를 재사용하되 두 기존
# harness(perf1b/perf1c)는 전혀 수정하지 않는다.
#
# 이 스크립트는 정확히 1회만 실행하도록 설계됐다 - 내부에 재시도/재검색/
# 스크롤/추가 페이지 열기/CAPTCHA 우회/context 재시작/--force 옵션이
# 전혀 없다. fake session/page/response는 전혀 쓰지 않으며, 실제
# production 함수를 그대로 호출한다. production 코드는 이 스크립트에서
# 절대 수정하지 않는다. settle_ms(5000)/quiet_period_ms(750)/
# poll_interval_ms(100)도 production 기본값 그대로 두고 변경하지 않는다.
import argparse
import json
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

ROOT_DIR = Path(__file__).resolve().parents[2]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

# 고정 단일 검색 조합(work order §3 그대로 - 확장/변경하지 않음).
JOB = {
    "tier": "single", "source_layer": "district",
    "query": "서울특별시 강동구 카페",
    "source_city": "서울특별시", "source_district": "강동구", "source_subregion": "강동구",
}
JOBS = [JOB]

PER_QUERY_LIMIT = 300
TARGET_COUNT = 300
SETTLE_MS = 5000
RETRY_COUNT = 0
SCROLL_COUNT = 0
PAGINATION_COUNT = 0

RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
RESULTS_DIR = Path(__file__).resolve().parent / "results" / "limit300b"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_XLSX_PATH = RESULTS_DIR / f"limit300b_single_query_{RUN_TIMESTAMP}.xlsx"
RESULT_JSON_PATH = RESULTS_DIR / f"limit300b_single_query_result_{RUN_TIMESTAMP}.json"
MARKER_PATH = RESULTS_DIR / "LIMIT300B_LIVE_STARTED.marker"


def _current_git_commit() -> str:
    try:
        out = subprocess.run(
            ["git", "rev-parse", "HEAD"], cwd=ROOT_DIR,
            capture_output=True, text=True, check=True,
        )
        return out.stdout.strip()
    except Exception as exc:
        return f"확인 불가: {type(exc).__name__}: {exc}"


def _print_config_summary():
    print(f"[limit300b] query_count=1, query={JOB['query']!r}")
    print(f"[limit300b] per_query_limit={PER_QUERY_LIMIT}, target_count={TARGET_COUNT}, settle_ms={SETTLE_MS}")
    print(f"[limit300b] retry_count={RETRY_COUNT}, scroll_count={SCROLL_COUNT}, pagination_count={PAGINATION_COUNT}")
    print("[limit300b] quiet_period_ms=750, poll_interval_ms=100(production 상수, harness가 값을 바꾸지 않음)")
    print(f"[limit300b] 결과 디렉터리: {RESULTS_DIR}")
    print(f"[limit300b] 실행 마커 경로: {MARKER_PATH}")
    print(f"[limit300b] Git commit: {_current_git_commit()}")
    print(f"[limit300b] Python executable: {sys.executable}")


def check_config() -> int:
    """--check-config: BrowserSession/Playwright를 전혀 시작하지 않고
    JOB 구성, 고정 실행 조건, production import 경로만 검증한다.
    live 마커/결과 파일을 생성하지 않는다."""
    print("[limit300b] ==== --check-config 모드(네이버 접속 없음) ====")
    _print_config_summary()

    from src.exporter import MERGED_COLUMNS, export_places_to_excel  # noqa: F401
    from src.pc.browser_session import BrowserSession  # noqa: F401
    from src.pc.config import DiagnosticConfig  # noqa: F401
    from src.pc.network_browser_collector import NetworkBrowserCollector, _SEARCH_URL_TEMPLATE  # noqa: F401
    from src.pc.network_list_scraper import dedup_rows  # noqa: F401
    from src.pc.network_pipeline import run_collection_plan  # noqa: F401

    print("[limit300b] production import 전부 성공")
    print(f"[limit300b] MARKER_PATH 존재 여부(참고용, 생성 안 함): {MARKER_PATH.exists()}")
    print("[limit300b] --check-config PASS - live를 실행하지 않았습니다")
    return 0


def main() -> int:
    from src.exporter import MERGED_COLUMNS, export_places_to_excel
    from src.pc.browser_session import BrowserSession
    from src.pc.config import DiagnosticConfig
    from src.pc.network_browser_collector import NetworkBrowserCollector, _SEARCH_URL_TEMPLATE
    from src.pc.network_list_scraper import dedup_rows as diagnostic_dedup_rows
    from src.pc.network_pipeline import run_collection_plan
    import openpyxl

    def _safe_session_factory():
        return BrowserSession(DiagnosticConfig.safe_default())

    if MARKER_PATH.exists():
        print(f"[limit300b] 실행 마커가 이미 존재합니다 - live 실행을 하지 않고 종료합니다: {MARKER_PATH}")
        return 1

    _print_config_summary()

    collected_at = datetime.now().strftime("%Y-%m-%d")

    timings: dict = {}
    per_query_log: list = []
    diagnostic_seen: set = set()

    def make_recording_collect_query(collector):
        def wrapper(job, per_query_limit):
            index = len(per_query_log) + 1
            search_url = _SEARCH_URL_TEMPLATE.format(query=quote(job["query"]))
            t0 = time.perf_counter()
            result = collector.collect_query(job, per_query_limit)
            t1 = time.perf_counter()

            rows = result.get("rows") or []
            newly_unique = diagnostic_dedup_rows(rows, diagnostic_seen)

            per_query_log.append({
                "index": index,
                "query": job["query"],
                "search_url": search_url,
                "recorded_per_query_limit": per_query_limit,
                "query_wall_seconds": t1 - t0,
                "candidate_response_count": result.get("candidate_response_count"),
                "raw_item_count": result.get("raw_item_count"),
                "local_unique_count": result.get("local_unique_count"),
                "returned_row_count": len(rows),
                "global_unique_added_diagnostic": len(newly_unique),
                "parse_error_count": result.get("parse_error_count"),
                "timeout": result.get("timeout"),
                "active_captcha_detected": result.get("active_captcha_detected"),
                "status_429_seen": result.get("status_429_seen"),
                "navigation_error": result.get("navigation_error"),
                "navigation_error_message": (result.get("navigation_error_message") or "")[:200],
                "adaptive_settle_wait_ms": result.get("adaptive_settle_wait_ms"),
                "adaptive_settle_early_exit": result.get("adaptive_settle_early_exit"),
            })
            print(f"[limit300b] query {index}/1 완료: {job['query']!r} recorded_per_query_limit={per_query_limit} "
                  f"candidate={result.get('candidate_response_count')} raw={result.get('raw_item_count')} "
                  f"local_unique={result.get('local_unique_count')} returned_rows={len(rows)} "
                  f"adaptive_wait_ms={result.get('adaptive_settle_wait_ms')} "
                  f"early_exit={result.get('adaptive_settle_early_exit')} "
                  f"wall={t1 - t0:.2f}s")
            return result

        return wrapper

    # 실행 마커를 BrowserSession/Playwright 시작 전, 원자적으로("x" 모드) 생성한다.
    marker_payload = {
        "started_at": datetime.now().isoformat(),
        "git_commit": _current_git_commit(),
        "python_executable": sys.executable,
        "harness_path": str(Path(__file__).resolve()),
        "query": JOB["query"],
        "per_query_limit": PER_QUERY_LIMIT,
        "target_count": TARGET_COUNT,
    }
    try:
        with open(MARKER_PATH, "x", encoding="utf-8") as f:
            f.write(json.dumps(marker_payload, ensure_ascii=False, indent=2))
    except FileExistsError:
        print(f"[limit300b] 실행 마커가 실행 직전 경합으로 이미 생성되었습니다 - 중단: {MARKER_PATH}")
        return 1
    print(f"[limit300b] 실행 마커 생성: {MARKER_PATH}")
    print("[limit300b] live 실행 시작(정확히 1회, 재시도/스크롤/페이지네이션 없음)")

    t_start = time.perf_counter()
    export_result_path = None
    export_error_message = ""
    exported = False
    exporter_call_count = 0

    def _counting_export_places_to_excel(merged_data, mobile_data, pc_data, output_path):
        nonlocal exporter_call_count
        exporter_call_count += 1
        return export_places_to_excel(merged_data, mobile_data, pc_data, output_path)

    with NetworkBrowserCollector(collected_at=collected_at, session_factory=_safe_session_factory, settle_ms=SETTLE_MS) as collector:
        t_session_ready = time.perf_counter()
        timings["session_ready_seconds"] = t_session_ready - t_start

        orchestrator_result = run_collection_plan(
            JOBS,
            per_query_limit=PER_QUERY_LIMIT,
            target_count=TARGET_COUNT,
            collected_at=collected_at,
            collect_query=make_recording_collect_query(collector),
        )
        t_orchestrator_done = time.perf_counter()
        timings["orchestrator_seconds"] = t_orchestrator_done - t_session_ready

    t_session_closed = time.perf_counter()
    timings["session_teardown_seconds"] = t_session_closed - t_orchestrator_done

    rows = orchestrator_result.get("rows") or []

    place_id_values = [str(r.get("place_id") or "").strip() for r in rows if r.get("place_id")]
    place_url_values = [str(r.get("플레이스 URL") or "").strip() for r in rows if r.get("플레이스 URL")]
    pre_export_dedup_check = {
        "rows_count": len(rows),
        "place_id_present_count": len(place_id_values),
        "place_id_missing_count": len(rows) - len(place_id_values),
        "place_id_duplicate_count": len(place_id_values) - len(set(place_id_values)),
        "place_url_present_count": len(place_url_values),
        "place_url_duplicate_count": len(place_url_values) - len(set(place_url_values)),
    }

    if rows:
        try:
            saved_path = _counting_export_places_to_excel(rows, [], [], str(OUTPUT_XLSX_PATH))
            exported = True
            export_result_path = str(saved_path or OUTPUT_XLSX_PATH)
        except Exception as exc:
            export_error_message = f"{type(exc).__name__}: {exc}"
            print(f"[limit300b] Excel 저장 실패: {export_error_message}")
        t_export_done = time.perf_counter()
        timings["export_seconds"] = t_export_done - t_session_closed
    else:
        timings["export_seconds"] = 0.0
        print("[limit300b] rows=0 - export_places_to_excel 호출하지 않음(파일 미생성)")

    excel_verification: dict = {}
    if exported and export_result_path:
        wb = openpyxl.load_workbook(export_result_path)
        merged_ws = wb["통합_결과"]
        headers = [cell.value for cell in merged_ws[1]]
        merged_data_rows = list(merged_ws.iter_rows(min_row=2, values_only=True))
        mobile_ws = wb["원본_모바일"]
        pc_ws = wb["원본_PC"]

        internal_fields = ("place_id", "source_city", "source_district", "source_subregion", "source_layer", "source_query")
        header_leak = [field for field in internal_fields if field in headers]

        url_idx = headers.index("플레이스 URL") if "플레이스 URL" in headers else None
        excel_urls = []
        if url_idx is not None:
            excel_urls = [row[url_idx] for row in merged_data_rows if row[url_idx] not in (None, "")]
        excel_url_duplicate_count = len(excel_urls) - len(set(excel_urls))

        new_open_idx = headers.index("새로오픈여부") if "새로오픈여부" in headers else None
        new_open_blank_count = None
        if new_open_idx is not None:
            new_open_blank_count = sum(1 for row in merged_data_rows if row[new_open_idx] in (None, ""))

        excel_verification = {
            "headers": headers,
            "headers_match_merged_columns": headers == MERGED_COLUMNS,
            "data_row_count": len(merged_data_rows),
            "data_row_count_matches_final_count": len(merged_data_rows) == orchestrator_result.get("final_count"),
            "internal_field_leak": header_leak,
            "excel_place_url_duplicate_count": excel_url_duplicate_count,
            "new_open_blank_count": new_open_blank_count,
            "mobile_sheet_max_row": mobile_ws.max_row,
            "pc_sheet_max_row": pc_ws.max_row,
        }

    t_end = time.perf_counter()
    timings["total_wall_seconds"] = t_end - t_start

    final_count = orchestrator_result.get("final_count") or 0
    available_ratio_percent = (final_count / 300 * 100) if TARGET_COUNT == 300 else None
    per_query_limit_unused = 300 - final_count if TARGET_COUNT == 300 else None

    # 판정(work order §14 그대로 - "네이버에 N개뿐" 같은 단정 표현은 쓰지 않는다).
    execution_contract_conditions = {
        "live_run_count_is_1": True,
        "retry_count_is_0": RETRY_COUNT == 0,
        "scroll_count_is_0": SCROLL_COUNT == 0,
        "pagination_count_is_0": PAGINATION_COUNT == 0,
        "total_job_count_is_1": orchestrator_result.get("executed_query_count", 0) + orchestrator_result.get("skipped_query_count", 0) == 1,
        "executed_query_count_is_1": orchestrator_result.get("executed_query_count") == 1,
        "recorded_per_query_limit_is_300": bool(per_query_log) and per_query_log[0]["recorded_per_query_limit"] == 300,
        "navigation_error_is_false": orchestrator_result.get("navigation_error") is False,
        "active_captcha_is_false": not any(entry.get("active_captcha_detected") for entry in per_query_log),
        "status_429_is_false": not any(entry.get("status_429_seen") for entry in per_query_log),
    }
    if final_count >= 1:
        execution_contract_conditions.update({
            "exported_is_true": exported is True,
            "exporter_call_count_is_1": exporter_call_count == 1,
            "excel_data_row_count_matches_final_count": excel_verification.get("data_row_count_matches_final_count") is True,
            "place_id_duplicate_count_is_0": pre_export_dedup_check["place_id_duplicate_count"] == 0,
            "place_url_duplicate_count_is_0": pre_export_dedup_check["place_url_duplicate_count"] == 0,
            "no_internal_field_leak": excel_verification.get("internal_field_leak") == [],
        })
    execution_contract_pass = all(execution_contract_conditions.values())

    single_combination_300_support = "미확인(final_count < 300)"
    if (
        final_count == 300
        and (orchestrator_result.get("before_trim_count") or 0) >= 300
        and orchestrator_result.get("stop_reason") == "target_reached"
        and excel_verification.get("data_row_count") == 300
    ):
        single_combination_300_support = "PASS(final_count=300, before_trim_count>=300, stop_reason=target_reached, Excel 300행 확인)"

    verdict = {
        "execution_contract_pass": execution_contract_pass,
        "execution_contract_conditions": execution_contract_conditions,
        "single_combination_300_support": single_combination_300_support,
        "interpretation_note": (
            "final_count는 '현재 production 경로에서 이번 1회 실행으로 확보한 개수'다. "
            "이 값이 300 미만이라고 해서 네이버 자체에 업체가 그만큼만 존재한다고 "
            "단정하지 않는다 - per_query_limit은 수집 결과에 적용되는 최대 허용값이지 "
            "수집량을 강제로 늘리는 값이 아니다."
        ),
    }

    summary = {
        "run_timestamp": RUN_TIMESTAMP,
        "git_commit": _current_git_commit(),
        "python_executable": sys.executable,
        "harness_path": str(Path(__file__).resolve()),
        "marker_path": str(MARKER_PATH),
        "query": JOB["query"],
        "total_job_count": 1,
        "per_query_limit": PER_QUERY_LIMIT,
        "target_count": TARGET_COUNT,
        "settle_ms": SETTLE_MS,
        "retry_count": RETRY_COUNT,
        "scroll_count": SCROLL_COUNT,
        "pagination_count": PAGINATION_COUNT,
        "executed_query_count": orchestrator_result.get("executed_query_count"),
        "skipped_query_count": orchestrator_result.get("skipped_query_count"),
        "before_trim_count": orchestrator_result.get("before_trim_count"),
        "final_count": final_count,
        "stop_reason": orchestrator_result.get("stop_reason"),
        "security_blocked": orchestrator_result.get("security_blocked"),
        "status_429_seen": orchestrator_result.get("status_429_seen"),
        "navigation_error": orchestrator_result.get("navigation_error"),
        "navigation_error_message": (orchestrator_result.get("navigation_error_message") or "")[:200],
        "exported": exported,
        "exporter_call_count": exporter_call_count,
        "export_path": export_result_path or "",
        "export_error": bool(export_error_message),
        "export_error_message": export_error_message,
        "available_ratio_percent": available_ratio_percent,
        "per_query_limit_unused": per_query_limit_unused,
        "pre_export_dedup_check": pre_export_dedup_check,
        "timings": timings,
        "per_query_log": per_query_log,
        "excel_verification": excel_verification,
        "verdict": verdict,
    }

    print("[limit300b] ==== 결과 요약 ====")
    for key, value in summary.items():
        if key in ("timings", "excel_verification", "per_query_log", "verdict"):
            continue
        print(f"[limit300b] {key}={value}")
    print(f"[limit300b] timings={timings}")
    print(f"[limit300b] excel_verification={excel_verification}")
    print(f"[limit300b] verdict={verdict}")

    RESULT_JSON_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[limit300b] 결과 JSON 저장: {RESULT_JSON_PATH}")

    if exported:
        print(f"[limit300b] Excel 저장 경로: {export_result_path}")

    print("[limit300b] 실행 종료(재실행하지 않음, 마커는 삭제하지 않음)")
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--check-config", action="store_true")
    args = parser.parse_args()
    if args.check_config:
        sys.exit(check_config())
    else:
        sys.exit(main())
