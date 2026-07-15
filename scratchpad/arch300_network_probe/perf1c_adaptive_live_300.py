# ARCH-300C PERF-1C: WIRE-4C와 동일한 24-job(Tier1 9 + Tier2 6 + Tier3 9)/300건
# 조건에서 PERF-1A 적응형 settle을 실제 Playwright + 네이버로 정확히 1회
# 실행해, 대규모 다중 쿼리에서도 정확성·안전 계약·성능 개선이 유지되는지
# 최종 실측하는 live harness. wire4c_product_live_300.py(Tier 큐 구성/진단
# 집계)와 perf1b_adaptive_live_50.py(실행 마커/exporter counting wrapper/
# 성능 비교 계산) 구조를 그대로 재사용하되 두 기존 파일은 수정하지 않는다.
#
# Tier 큐 구성 근거는 wire4c_product_live_300.py의 것과 동일하다(PoC-7
# 실제 24개 큐, "베이커리카페" 포함) - 이번 work order 본문도 이미 9+6+9
# 24개로 정확히 일치하므로 큐 구성에 관한 추가 대조/수정은 필요하지 않았다.
#
# 이 스크립트는 정확히 1회만 실행하도록 설계됐다 - 내부에 재시도/재검색/
# CAPTCHA 우회/context 재시작/--force 옵션이 전혀 없다. fake session/page/
# response는 전혀 쓰지 않으며, 실제 production 함수를 그대로 호출한다.
# production 코드는 이 스크립트에서 절대 수정하지 않는다.
import argparse
import json
import statistics
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

ROOT_DIR = Path(__file__).resolve().parents[2]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

_COMMON_META = {"source_city": "서울특별시", "source_district": "강동구"}

# Tier1: 법정동(legal_dong) - wire4c_product_live_300.py와 순서까지 동일.
_TIER1_LEGAL_DONGS = ["천호동", "성내동", "길동", "암사동", "명일동", "고덕동", "상일동", "둔촌동", "강일동"]
# Tier2: 역/상권(landmark) - 동일 출처.
_TIER2_LANDMARKS = ["천호역", "강동역", "둔촌동역", "암사역", "고덕역", "명일역"]
# Tier3: 세부업종(vertical) - "베이커리카페" 포함 9개(work order 본문과 이미 일치).
_TIER3_DONGS = ["천호동", "성내동", "길동"]
_TIER3_SUBCATEGORIES = ["디저트카페", "브런치카페", "베이커리카페"]

JOBS: list = []
for dong in _TIER1_LEGAL_DONGS:
    JOBS.append({
        "tier": "tier1", "source_layer": "legal_dong",
        "source_subregion": dong, "query": f"서울특별시 강동구 {dong} 카페", **_COMMON_META,
    })
for landmark in _TIER2_LANDMARKS:
    JOBS.append({
        "tier": "tier2", "source_layer": "landmark",
        "source_subregion": landmark, "query": f"서울특별시 강동구 {landmark} 카페", **_COMMON_META,
    })
for dong in _TIER3_DONGS:
    for subcategory in _TIER3_SUBCATEGORIES:
        JOBS.append({
            "tier": "tier3", "source_layer": "vertical",
            "source_subregion": dong, "query": f"서울특별시 강동구 {dong} {subcategory}", **_COMMON_META,
        })

assert len(JOBS) == 24, f"큐 구성 오류: 예상 24개, 실제 {len(JOBS)}개"

PER_QUERY_LIMIT = 20
TARGET_COUNT = 300
SETTLE_MS = 5000

# WIRE-4C 실측 기준선(고정 5초 settle, 동일 24-job/300건 조건).
WIRE4C_BASELINE = {
    "total_wall_seconds": 100.72,
    "orchestrator_seconds": 99.71,
    "avg_query_wall_seconds": 5.87,
    "rows_per_second": 2.98,
    "seconds_per_final_row": 0.336,
}
# PERF-1B 실측(참고 비교용, 판정 기준으로 쓰지 않음).
PERF1B_OBSERVED = {
    "avg_query_wall_seconds": 2.580,
    "adaptive_wait_avg_ms": 1600.0,
}

RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
RESULTS_DIR = Path(__file__).resolve().parent / "results" / "perf1c"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_XLSX_PATH = RESULTS_DIR / f"perf1c_adaptive_live_300_{RUN_TIMESTAMP}.xlsx"
RESULT_JSON_PATH = RESULTS_DIR / f"perf1c_adaptive_live_300_result_{RUN_TIMESTAMP}.json"
MARKER_PATH = RESULTS_DIR / "PERF1C_LIVE_STARTED.marker"


def _current_git_commit() -> str:
    try:
        out = subprocess.run(
            ["git", "rev-parse", "HEAD"], cwd=ROOT_DIR,
            capture_output=True, text=True, check=True,
        )
        return out.stdout.strip()
    except Exception as exc:
        return f"확인 불가: {type(exc).__name__}: {exc}"


def _percentile95(values: list) -> float | None:
    """단순 최근접 순위법(nearest-rank): sorted(values)[ceil(0.95*n) - 1].
    표본이 매우 작을 때(n<20)는 근사치일 뿐이라는 점을 결과 JSON에 함께 기록한다."""
    if not values:
        return None
    ordered = sorted(values)
    n = len(ordered)
    rank = max(1, -(-95 * n // 100))  # ceil(0.95 * n), 최소 1
    return ordered[min(rank, n) - 1]


def _print_config_summary():
    tier1 = sum(1 for j in JOBS if j["tier"] == "tier1")
    tier2 = sum(1 for j in JOBS if j["tier"] == "tier2")
    tier3 = sum(1 for j in JOBS if j["tier"] == "tier3")
    print(f"[perf1c] query 24개 확인: tier1={tier1}, tier2={tier2}, tier3={tier3}")
    print(f"[perf1c] tier 순서: {[j['tier'] for j in JOBS]}")
    print(f"[perf1c] per_query_limit={PER_QUERY_LIMIT}, target_count={TARGET_COUNT}, settle_ms={SETTLE_MS}")
    print("[perf1c] quiet_period_ms=750, poll_interval_ms=100(production 상수, harness가 값을 바꾸지 않음)")
    print("[perf1c] retry=0(자동 재시도/수동 재실행 없음)")
    print(f"[perf1c] 결과 디렉터리: {RESULTS_DIR}")
    print(f"[perf1c] 실행 마커 경로: {MARKER_PATH}")
    print(f"[perf1c] Git commit: {_current_git_commit()}")
    print(f"[perf1c] Python executable: {sys.executable}")


def check_config() -> int:
    """--check-config: BrowserSession/Playwright를 전혀 시작하지 않고
    JOBS 구성, 고정 실행 조건, production import 경로만 검증한다.
    live 마커/결과 파일을 생성하지 않는다."""
    print("[perf1c] ==== --check-config 모드(네이버 접속 없음) ====")
    _print_config_summary()

    # production import가 실제로 가능한지만 확인한다(인스턴스화/실행은 하지 않음).
    from src.exporter import MERGED_COLUMNS, export_places_to_excel  # noqa: F401
    from src.pc.browser_session import BrowserSession  # noqa: F401
    from src.pc.config import DiagnosticConfig  # noqa: F401
    from src.pc.network_browser_collector import NetworkBrowserCollector, _SEARCH_URL_TEMPLATE  # noqa: F401
    from src.pc.network_list_scraper import classify_query_efficiency, dedup_rows  # noqa: F401
    from src.pc.network_pipeline import run_collection_plan  # noqa: F401

    print("[perf1c] production import 전부 성공")
    print(f"[perf1c] MARKER_PATH 존재 여부(참고용, 생성 안 함): {MARKER_PATH.exists()}")
    print("[perf1c] --check-config PASS - live를 실행하지 않았습니다")
    return 0


def main() -> int:
    from src.exporter import MERGED_COLUMNS, export_places_to_excel
    from src.pc.browser_session import BrowserSession
    from src.pc.config import DiagnosticConfig
    from src.pc.network_browser_collector import NetworkBrowserCollector, _SEARCH_URL_TEMPLATE
    from src.pc.network_list_scraper import classify_query_efficiency, dedup_rows as diagnostic_dedup_rows
    from src.pc.network_pipeline import run_collection_plan
    import openpyxl

    def _safe_session_factory():
        return BrowserSession(DiagnosticConfig.safe_default())

    if MARKER_PATH.exists():
        print(f"[perf1c] 실행 마커가 이미 존재합니다 - live 실행을 하지 않고 종료합니다: {MARKER_PATH}")
        return 1

    _print_config_summary()

    collected_at = datetime.now().strftime("%Y-%m-%d")

    timings: dict = {}
    per_query_log: list = []
    diagnostic_seen: set = set()
    cumulative_unique = 0

    def make_recording_collect_query(collector):
        def wrapper(job, per_query_limit):
            nonlocal cumulative_unique
            index = len(per_query_log) + 1
            search_url = _SEARCH_URL_TEMPLATE.format(query=quote(job["query"]))
            t0 = time.perf_counter()
            result = collector.collect_query(job, per_query_limit)
            t1 = time.perf_counter()

            rows = result.get("rows") or []
            newly_unique = diagnostic_dedup_rows(rows, diagnostic_seen)
            cumulative_unique += len(newly_unique)
            raw_item_count = result.get("raw_item_count") or 0
            efficiency = classify_query_efficiency(raw_item_count, len(newly_unique))

            per_query_log.append({
                "index": index,
                "tier": job.get("tier"),
                "source_layer": job.get("source_layer"),
                "source_subregion": job.get("source_subregion"),
                "query": job["query"],
                "search_url": search_url,
                "query_wall_seconds": t1 - t0,
                "candidate_response_count": result.get("candidate_response_count"),
                "raw_item_count": raw_item_count,
                "local_unique_count": result.get("local_unique_count"),
                "returned_row_count": len(rows),
                "global_unique_added_diagnostic": len(newly_unique),
                "global_duplicate_count_diagnostic": len(rows) - len(newly_unique),
                "cumulative_unique_diagnostic": cumulative_unique,
                "efficiency_ratio_diagnostic": efficiency["efficiency_ratio"],
                "parse_error_count": result.get("parse_error_count"),
                "timeout": result.get("timeout"),
                "active_captcha_detected": result.get("active_captcha_detected"),
                "status_429_seen": result.get("status_429_seen"),
                "navigation_error": result.get("navigation_error"),
                "navigation_error_message": (result.get("navigation_error_message") or "")[:200],
                "adaptive_settle_wait_ms": result.get("adaptive_settle_wait_ms"),
                "adaptive_settle_early_exit": result.get("adaptive_settle_early_exit"),
            })
            print(f"[perf1c] query {index}/{len(JOBS)} [{job.get('tier')}] 완료: {job['query']!r} "
                  f"candidate={result.get('candidate_response_count')} raw={raw_item_count} "
                  f"local_unique={result.get('local_unique_count')} returned_rows={len(rows)} "
                  f"adaptive_wait_ms={result.get('adaptive_settle_wait_ms')} "
                  f"early_exit={result.get('adaptive_settle_early_exit')} "
                  f"wall={t1 - t0:.2f}s")
            return result

        return wrapper

    # 실행 마커를 BrowserSession/Playwright 시작 전, 원자적으로("x" 모드) 생성한다.
    marker_payload = {
        "started_at": datetime.now().isoformat(),
        "git_commit": _current_git_commit(),
        "python_executable": sys.executable,
        "harness_path": str(Path(__file__).resolve()),
        "total_job_count": len(JOBS),
        "per_query_limit": PER_QUERY_LIMIT,
        "target_count": TARGET_COUNT,
    }
    try:
        with open(MARKER_PATH, "x", encoding="utf-8") as f:
            f.write(json.dumps(marker_payload, ensure_ascii=False, indent=2))
    except FileExistsError:
        print(f"[perf1c] 실행 마커가 실행 직전 경합으로 이미 생성되었습니다 - 중단: {MARKER_PATH}")
        return 1
    print(f"[perf1c] 실행 마커 생성: {MARKER_PATH}")
    print("[perf1c] live 실행 시작(정확히 1회, 재시도 없음)")

    t_start = time.perf_counter()
    export_result_path = None
    export_error_message = ""
    exported = False
    exporter_call_count = 0

    def _counting_export_places_to_excel(merged_data, mobile_data, pc_data, output_path):
        nonlocal exporter_call_count
        exporter_call_count += 1
        return export_places_to_excel(merged_data, mobile_data, pc_data, output_path)

    with NetworkBrowserCollector(collected_at=collected_at, session_factory=_safe_session_factory, settle_ms=SETTLE_MS) as collector:
        t_session_ready = time.perf_counter()
        timings["session_ready_seconds"] = t_session_ready - t_start

        orchestrator_result = run_collection_plan(
            JOBS,
            per_query_limit=PER_QUERY_LIMIT,
            target_count=TARGET_COUNT,
            collected_at=collected_at,
            collect_query=make_recording_collect_query(collector),
        )
        t_orchestrator_done = time.perf_counter()
        timings["orchestrator_seconds"] = t_orchestrator_done - t_session_ready

    t_session_closed = time.perf_counter()
    timings["session_teardown_seconds"] = t_session_closed - t_orchestrator_done

    rows = orchestrator_result.get("rows") or []

    place_id_values = [str(r.get("place_id") or "").strip() for r in rows if r.get("place_id")]
    place_url_values = [str(r.get("플레이스 URL") or "").strip() for r in rows if r.get("플레이스 URL")]
    pre_export_dedup_check = {
        "rows_count": len(rows),
        "place_id_present_count": len(place_id_values),
        "place_id_missing_count": len(rows) - len(place_id_values),
        "place_id_duplicate_count": len(place_id_values) - len(set(place_id_values)),
        "place_url_present_count": len(place_url_values),
        "place_url_duplicate_count": len(place_url_values) - len(set(place_url_values)),
    }

    if rows:
        try:
            saved_path = _counting_export_places_to_excel(rows, [], [], str(OUTPUT_XLSX_PATH))
            exported = True
            export_result_path = str(saved_path or OUTPUT_XLSX_PATH)
        except Exception as exc:
            export_error_message = f"{type(exc).__name__}: {exc}"
            print(f"[perf1c] Excel 저장 실패: {export_error_message}")
        t_export_done = time.perf_counter()
        timings["export_seconds"] = t_export_done - t_session_closed
    else:
        timings["export_seconds"] = 0.0
        print("[perf1c] rows=0 - export_places_to_excel 호출하지 않음(파일 미생성)")

    excel_verification: dict = {}
    if exported and export_result_path:
        wb = openpyxl.load_workbook(export_result_path)
        merged_ws = wb["통합_결과"]
        headers = [cell.value for cell in merged_ws[1]]
        merged_data_rows = list(merged_ws.iter_rows(min_row=2, values_only=True))
        mobile_ws = wb["원본_모바일"]
        pc_ws = wb["원본_PC"]

        internal_fields = ("place_id", "source_city", "source_district", "source_subregion", "source_layer", "source_query")
        header_leak = [field for field in internal_fields if field in headers]

        url_idx = headers.index("플레이스 URL") if "플레이스 URL" in headers else None
        excel_urls = []
        if url_idx is not None:
            excel_urls = [row[url_idx] for row in merged_data_rows if row[url_idx] not in (None, "")]
        excel_url_duplicate_count = len(excel_urls) - len(set(excel_urls))

        new_open_idx = headers.index("새로오픈여부") if "새로오픈여부" in headers else None
        new_open_blank_count = None
        if new_open_idx is not None:
            new_open_blank_count = sum(1 for row in merged_data_rows if row[new_open_idx] in (None, ""))

        def _count_filled(col_name: str):
            if col_name not in headers:
                return None
            idx = headers.index(col_name)
            return sum(1 for row in merged_data_rows if row[idx] not in (None, ""))

        excel_verification = {
            "headers": headers,
            "headers_match_merged_columns": headers == MERGED_COLUMNS,
            "data_row_count": len(merged_data_rows),
            "data_row_count_matches_final_count": len(merged_data_rows) == orchestrator_result.get("final_count"),
            "internal_field_leak": header_leak,
            "excel_place_url_duplicate_count": excel_url_duplicate_count,
            "new_open_blank_count": new_open_blank_count,
            "mobile_sheet_max_row": mobile_ws.max_row,
            "pc_sheet_max_row": pc_ws.max_row,
            "field_fill_counts": {col: _count_filled(col) for col in MERGED_COLUMNS},
        }

    t_end = time.perf_counter()
    timings["total_wall_seconds"] = t_end - t_start
    executed_count = len(per_query_log)
    timings["avg_query_wall_seconds"] = (
        sum(entry["query_wall_seconds"] for entry in per_query_log) / executed_count if executed_count else 0.0
    )
    final_count = orchestrator_result.get("final_count") or 0
    timings["rows_per_second"] = (final_count / timings["total_wall_seconds"]) if timings["total_wall_seconds"] else 0.0
    timings["seconds_per_final_row"] = (timings["total_wall_seconds"] / final_count) if final_count else None

    total_raw_items = sum((entry["raw_item_count"] or 0) for entry in per_query_log)
    total_returned_rows = sum(entry["returned_row_count"] for entry in per_query_log)
    before_trim_count = orchestrator_result.get("before_trim_count") or 0
    global_duplicates_removed = total_returned_rows - before_trim_count
    global_dedup_rate = (global_duplicates_removed / total_returned_rows) if total_returned_rows else 0.0
    diagnostic_matches_orchestrator = cumulative_unique == before_trim_count
    parse_error_count_total = sum((entry["parse_error_count"] or 0) for entry in per_query_log)
    timeout_count = sum(1 for entry in per_query_log if entry["timeout"])

    tier_summary: dict = {}
    for entry in per_query_log:
        tier = entry["tier"] or "unknown"
        bucket = tier_summary.setdefault(tier, {
            "executed_count": 0, "raw_total": 0, "local_unique_total": 0,
            "returned_rows_total": 0, "global_unique_added_total": 0,
            "adaptive_wait_ms_total": 0, "query_wall_seconds_total": 0.0,
        })
        bucket["executed_count"] += 1
        bucket["raw_total"] += entry["raw_item_count"] or 0
        bucket["local_unique_total"] += entry["local_unique_count"] or 0
        bucket["returned_rows_total"] += entry["returned_row_count"]
        bucket["global_unique_added_total"] += entry["global_unique_added_diagnostic"]
        bucket["adaptive_wait_ms_total"] += entry.get("adaptive_settle_wait_ms") or 0
        bucket["query_wall_seconds_total"] += entry["query_wall_seconds"]

    # ---- adaptive 파생 지표 ----
    valid_wait_values = []
    adaptive_metadata_missing_count = 0
    for entry in per_query_log:
        wait_ms = entry.get("adaptive_settle_wait_ms")
        early_exit = entry.get("adaptive_settle_early_exit")
        if not isinstance(wait_ms, (int, float)) or not isinstance(early_exit, bool):
            adaptive_metadata_missing_count += 1
            continue
        valid_wait_values.append(wait_ms)

    early_exit_query_count = sum(
        1 for entry in per_query_log if entry.get("adaptive_settle_early_exit") is True
    )
    hard_cap_query_count = sum(
        1 for entry in per_query_log
        if entry.get("adaptive_settle_early_exit") is False
        and isinstance(entry.get("adaptive_settle_wait_ms"), (int, float))
        and entry.get("adaptive_settle_wait_ms") >= SETTLE_MS
    )
    adaptive_wait_sum_ms = sum(valid_wait_values)
    fixed_wait_equivalent_ms = executed_count * 5000
    wait_reduction_ms = fixed_wait_equivalent_ms - adaptive_wait_sum_ms
    wait_reduction_percent = (
        (wait_reduction_ms / fixed_wait_equivalent_ms * 100) if fixed_wait_equivalent_ms else 0.0
    )

    adaptive_wait_distribution = {
        "sample_count": len(valid_wait_values),
        "minimum_ms": min(valid_wait_values) if valid_wait_values else None,
        "median_ms": statistics.median(valid_wait_values) if valid_wait_values else None,
        "average_ms": (sum(valid_wait_values) / len(valid_wait_values)) if valid_wait_values else None,
        "p95_ms": _percentile95(valid_wait_values),
        "maximum_ms": max(valid_wait_values) if valid_wait_values else None,
        "p95_method": "nearest_rank(ceil(0.95*n)) - 표본 수가 작으면(n<20) 근사치",
    }

    # ---- 성능 비교(WIRE-4C 기준선 대비) ----
    total_wall_improvement_seconds = WIRE4C_BASELINE["total_wall_seconds"] - timings["total_wall_seconds"]
    total_wall_improvement_percent = (
        total_wall_improvement_seconds / WIRE4C_BASELINE["total_wall_seconds"] * 100
    )
    orchestrator_improvement_seconds = WIRE4C_BASELINE["orchestrator_seconds"] - timings["orchestrator_seconds"]
    orchestrator_improvement_percent = (
        orchestrator_improvement_seconds / WIRE4C_BASELINE["orchestrator_seconds"] * 100
    )
    avg_query_improvement_seconds = WIRE4C_BASELINE["avg_query_wall_seconds"] - timings["avg_query_wall_seconds"]
    avg_query_improvement_percent = (
        avg_query_improvement_seconds / WIRE4C_BASELINE["avg_query_wall_seconds"] * 100
    )

    performance_comparison = {
        "wire4c_baseline": WIRE4C_BASELINE,
        "perf1b_observed_for_reference_only": PERF1B_OBSERVED,
        "total_wall_improvement_seconds": total_wall_improvement_seconds,
        "total_wall_improvement_percent": total_wall_improvement_percent,
        "orchestrator_improvement_seconds": orchestrator_improvement_seconds,
        "orchestrator_improvement_percent": orchestrator_improvement_percent,
        "avg_query_improvement_seconds": avg_query_improvement_seconds,
        "avg_query_improvement_percent": avg_query_improvement_percent,
        "adaptive_wait_sum_ms": adaptive_wait_sum_ms,
        "fixed_wait_equivalent_ms": fixed_wait_equivalent_ms,
        "wait_reduction_ms": wait_reduction_ms,
        "wait_reduction_percent": wait_reduction_percent,
        "early_exit_query_count": early_exit_query_count,
        "hard_cap_query_count": hard_cap_query_count,
        "adaptive_metadata_missing_count": adaptive_metadata_missing_count,
        "note": "WIRE-4C와 PERF-1C는 다른 시점의 네이버 환경이므로 네트워크·서비스 상태 "
                "차이가 개선폭에 포함될 수 있다. 이번 1회 결과를 일반적인 속도 보장으로 "
                "표현하지 않는다.",
    }

    summary = {
        "run_timestamp": RUN_TIMESTAMP,
        "git_commit": _current_git_commit(),
        "python_executable": sys.executable,
        "harness_path": str(Path(__file__).resolve()),
        "marker_path": str(MARKER_PATH),
        "total_job_count": len(JOBS),
        "per_query_limit": PER_QUERY_LIMIT,
        "target_count": TARGET_COUNT,
        "settle_ms": SETTLE_MS,
        "retries": 0,
        "executed_query_count": orchestrator_result.get("executed_query_count"),
        "skipped_query_count": orchestrator_result.get("skipped_query_count"),
        "before_trim_count": orchestrator_result.get("before_trim_count"),
        "final_count": orchestrator_result.get("final_count"),
        "stop_reason": orchestrator_result.get("stop_reason"),
        "security_blocked": orchestrator_result.get("security_blocked"),
        "status_429_seen": orchestrator_result.get("status_429_seen"),
        "navigation_error": orchestrator_result.get("navigation_error"),
        "navigation_error_message": (orchestrator_result.get("navigation_error_message") or "")[:200],
        "parse_error_count_total": parse_error_count_total,
        "timeout_count": timeout_count,
        "exported": exported,
        "exporter_call_count": exporter_call_count,
        "export_path": export_result_path or "",
        "export_error": bool(export_error_message),
        "export_error_message": export_error_message,
        "total_raw_items": total_raw_items,
        "total_local_unique": sum((entry["local_unique_count"] or 0) for entry in per_query_log),
        "total_returned_rows_after_local_dedup_and_cap": total_returned_rows,
        "global_duplicates_removed": global_duplicates_removed,
        "global_dedup_rate": global_dedup_rate,
        "diagnostic_cumulative_unique_matches_orchestrator_before_trim": diagnostic_matches_orchestrator,
        "pre_export_dedup_check": pre_export_dedup_check,
        "tier_summary": tier_summary,
        "timings": timings,
        "adaptive_wait_distribution": adaptive_wait_distribution,
        "per_query_log": per_query_log,
        "excel_verification": excel_verification,
        "performance_comparison": performance_comparison,
    }

    print("[perf1c] ==== 결과 요약 ====")
    for key, value in summary.items():
        if key in ("timings", "excel_verification", "per_query_log", "tier_summary",
                    "performance_comparison", "adaptive_wait_distribution"):
            continue
        print(f"[perf1c] {key}={value}")
    print(f"[perf1c] tier_summary={tier_summary}")
    print(f"[perf1c] timings={timings}")
    print(f"[perf1c] adaptive_wait_distribution={adaptive_wait_distribution}")
    print(f"[perf1c] excel_verification={excel_verification}")
    print(f"[perf1c] performance_comparison={performance_comparison}")

    RESULT_JSON_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[perf1c] 결과 JSON 저장: {RESULT_JSON_PATH}")

    if exported:
        print(f"[perf1c] Excel 저장 경로: {export_result_path}")

    print("[perf1c] 실행 종료(재실행하지 않음, 마커는 삭제하지 않음)")
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--check-config", action="store_true")
    args = parser.parse_args()
    if args.check_config:
        sys.exit(check_config())
    else:
        sys.exit(main())
