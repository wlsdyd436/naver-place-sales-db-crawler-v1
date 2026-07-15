# ARCH-300C WIRE-4A: 실제 Playwright + 네이버로 production 코드 경로(NetworkBrowserCollector
# -> collect_network_query -> run_collection_plan -> export_places_to_excel)를 쿼리 1개,
# 목표 10건으로 단발성 실행하는 live harness.
#
# 이 스크립트는 정확히 1회만 실행하도록 설계됐다 - 내부에 재시도/재검색/CAPTCHA 우회/
# context 재시작 로직이 전혀 없다. fake session/page/response는 전혀 쓰지 않으며,
# src/pc/network_browser_collector.py, src/pc/network_pipeline.py, src/exporter.py의
# 실제 함수를 그대로 호출한다. production 코드는 이 스크립트에서 절대 수정하지 않는다.
#
# 타이밍 계측은 production 함수 호출 앞뒤에서만 wall-clock을 재는 방식으로 수행한다.
# collect_query() 내부(page 생성/goto/settle/파싱/CAPTCHA probe)는 production 코드를
# 건드리거나 Playwright page 객체를 monkeypatch하지 않기 위해 세부 구간으로 쪼개지
# 않고 "쿼리 실행 전체 소요 시간"으로 하나로 기록한다.
from datetime import datetime
from pathlib import Path
from urllib.parse import quote
import json
import sys
import time

import openpyxl

ROOT_DIR = Path(__file__).resolve().parents[2]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.exporter import MERGED_COLUMNS, export_places_to_excel
from src.pc.browser_session import BrowserSession
from src.pc.config import DiagnosticConfig
from src.pc.network_browser_collector import NetworkBrowserCollector, _SEARCH_URL_TEMPLATE
from src.pc.network_pipeline import run_collection_plan

JOB = {
    "query": "서울특별시 강동구 천호동 카페",
    "source_city": "서울특별시",
    "source_district": "강동구",
    "source_subregion": "천호동",
    "source_layer": "legal_dong",
}
PER_QUERY_LIMIT = 10
TARGET_COUNT = 10

RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_XLSX_PATH = ROOT_DIR / "output" / f"wire4a_live_10_{RUN_TIMESTAMP}.xlsx"
RESULT_JSON_PATH = Path(__file__).resolve().parent / f"wire4a_live_10_result_{RUN_TIMESTAMP}.json"


def _safe_session_factory():
    # "기본 BrowserSession 경로"를 명시적으로 사용한다(고객용 안전 모드 -
    # visible=False/capture_artifacts=False/keep_open_on_error=False).
    return BrowserSession(DiagnosticConfig.safe_default())


def main() -> int:
    collected_at = datetime.now().strftime("%Y-%m-%d")
    expected_search_url = _SEARCH_URL_TEMPLATE.format(query=quote(JOB["query"]))

    timings: dict = {}
    per_query_results: list = []

    def recording_collect_query(collector):
        """collector.collect_query를 그대로 호출하되(동작 변경 없음), 반환값을
        harness 보고용으로만 기록한다 - production 함수 자체는 수정하지 않는다."""

        def wrapper(job, per_query_limit):
            t0 = time.perf_counter()
            per_query_result = collector.collect_query(job, per_query_limit)
            t1 = time.perf_counter()
            per_query_result = dict(per_query_result)
            per_query_result["_query_wall_seconds"] = t1 - t0
            per_query_results.append(per_query_result)
            return per_query_result

        return wrapper

    print(f"[wire4a] query={JOB['query']!r}")
    print(f"[wire4a] expected_search_url={expected_search_url}")
    print(f"[wire4a] per_query_limit={PER_QUERY_LIMIT}, target_count={TARGET_COUNT}")
    print("[wire4a] live 실행 시작(정확히 1회, 재시도 없음)")

    t_start = time.perf_counter()
    orchestrator_result = None
    export_result_path = None
    export_error_message = ""
    exported = False

    with NetworkBrowserCollector(collected_at=collected_at, session_factory=_safe_session_factory) as collector:
        t_session_ready = time.perf_counter()
        timings["session_ready_seconds"] = t_session_ready - t_start

        orchestrator_result = run_collection_plan(
            [JOB],
            per_query_limit=PER_QUERY_LIMIT,
            target_count=TARGET_COUNT,
            collected_at=collected_at,
            collect_query=recording_collect_query(collector),
        )
        t_query_done = time.perf_counter()
        timings["orchestrator_seconds"] = t_query_done - t_session_ready

    t_session_closed = time.perf_counter()
    timings["session_teardown_seconds"] = t_session_closed - t_query_done

    rows = orchestrator_result.get("rows") or []
    if rows:
        try:
            saved_path = export_places_to_excel(rows, [], [], str(OUTPUT_XLSX_PATH))
            exported = True
            export_result_path = str(saved_path or OUTPUT_XLSX_PATH)
        except Exception as exc:
            export_error_message = f"{type(exc).__name__}: {exc}"
            print(f"[wire4a] Excel 저장 실패: {export_error_message}")
        t_export_done = time.perf_counter()
        timings["export_seconds"] = t_export_done - t_session_closed
    else:
        timings["export_seconds"] = 0.0
        print("[wire4a] rows=0 - export_places_to_excel 호출하지 않음(파일 미생성)")

    excel_verification: dict = {}
    if exported and export_result_path:
        wb = openpyxl.load_workbook(export_result_path)
        merged_ws = wb["통합_결과"]
        headers = [cell.value for cell in merged_ws[1]]
        merged_data_rows = list(merged_ws.iter_rows(min_row=2, values_only=True))
        mobile_ws = wb["원본_모바일"]
        pc_ws = wb["원본_PC"]

        internal_fields = ("place_id", "source_city", "source_district", "source_subregion", "source_layer", "source_query")
        header_leak = [field for field in internal_fields if field in headers]

        def _count_filled(col_name: str):
            if col_name not in headers:
                return None
            idx = headers.index(col_name)
            return sum(1 for row in merged_data_rows if row[idx] not in (None, ""))

        excel_verification = {
            "headers": headers,
            "headers_match_merged_columns": headers == MERGED_COLUMNS,
            "data_row_count": len(merged_data_rows),
            "data_row_count_matches_final_count": len(merged_data_rows) == orchestrator_result.get("final_count"),
            "internal_field_leak": header_leak,
            "mobile_sheet_max_row": mobile_ws.max_row,
            "pc_sheet_max_row": pc_ws.max_row,
            "field_fill_counts": {col: _count_filled(col) for col in MERGED_COLUMNS},
        }

    t_end = time.perf_counter()
    timings["total_wall_seconds"] = t_end - t_start

    per_query = per_query_results[0] if per_query_results else {}

    summary = {
        "query": JOB["query"],
        "expected_search_url": expected_search_url,
        "per_query_limit": PER_QUERY_LIMIT,
        "target_count": TARGET_COUNT,
        "candidate_response_count": per_query.get("candidate_response_count"),
        "raw_item_count": per_query.get("raw_item_count"),
        "local_unique_count": per_query.get("local_unique_count"),
        "parse_error_count": per_query.get("parse_error_count"),
        "timeout": per_query.get("timeout"),
        "active_captcha_detected": per_query.get("active_captcha_detected"),
        "status_429_seen": per_query.get("status_429_seen"),
        "navigation_error": per_query.get("navigation_error"),
        "navigation_error_message": (per_query.get("navigation_error_message") or "")[:200],
        "query_wall_seconds": per_query.get("_query_wall_seconds"),
        "executed_query_count": orchestrator_result.get("executed_query_count"),
        "skipped_query_count": orchestrator_result.get("skipped_query_count"),
        "before_trim_count": orchestrator_result.get("before_trim_count"),
        "final_count": orchestrator_result.get("final_count"),
        "stop_reason": orchestrator_result.get("stop_reason"),
        "security_blocked": orchestrator_result.get("security_blocked"),
        "status_429_seen_orchestrator": orchestrator_result.get("status_429_seen"),
        "exported": exported,
        "export_path": export_result_path or "",
        "export_error": bool(export_error_message),
        "export_error_message": export_error_message,
        "timings": timings,
        "excel_verification": excel_verification,
        "retries": 0,
    }

    print("[wire4a] ==== 결과 요약 ====")
    for key, value in summary.items():
        if key in ("timings", "excel_verification"):
            continue
        print(f"[wire4a] {key}={value}")
    print(f"[wire4a] timings={timings}")
    print(f"[wire4a] excel_verification={excel_verification}")

    RESULT_JSON_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[wire4a] 결과 JSON 저장: {RESULT_JSON_PATH}")

    if exported:
        print(f"[wire4a] Excel 저장 경로: {export_result_path}")

    print("[wire4a] 실행 종료(재실행하지 않음)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
