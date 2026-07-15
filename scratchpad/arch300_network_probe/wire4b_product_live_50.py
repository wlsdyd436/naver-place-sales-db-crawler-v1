# ARCH-300C WIRE-4B: 실제 Playwright + 네이버로 production 코드 경로(NetworkBrowserCollector
# -> collect_network_query -> run_collection_plan -> export_places_to_excel)를 검색
# 조합 4개, 목표 50건으로 단발성 실행하는 live harness. WIRE-4A harness 구조를
# 그대로 재사용하되, job을 4개로 늘리고 쿼리별 진단 기록 + 진단 전용 global
# unique_added 계산을 추가했다.
#
# 이 스크립트는 정확히 1회만 실행하도록 설계됐다 - 내부에 재시도/재검색/CAPTCHA 우회/
# context 재시작 로직이 전혀 없다. fake session/page/response는 전혀 쓰지 않으며,
# src/pc/network_browser_collector.py, src/pc/network_pipeline.py, src/exporter.py의
# 실제 함수를 그대로 호출한다. production 코드는 이 스크립트에서 절대 수정하지 않는다.
#
# global unique_added 진단값은 run_collection_plan이 실제로 쓰는 dedup_rows를
# "읽기 전용으로" 재사용하되, run_collection_plan 내부의 실제 seen set과는 완전히
# 분리된 별도의 진단 전용 seen set에만 적용한다 - run_collection_plan에는 이
# 진단 seen을 전달하지 않으므로 실제 오케스트레이션 동작에는 전혀 개입하지 않는다.
#
# 타이밍 계측은 production 함수 호출 앞뒤에서만 wall-clock을 재는 방식으로 수행한다.
# collect_query() 내부(page 생성/goto/settle=5000ms 고정/파싱/CAPTCHA probe)는
# production 코드나 Playwright page 객체를 건드리지 않기 위해 세부 구간으로
# 쪼개지 않고 "쿼리 실행 전체 소요 시간"으로 하나로 기록한다.
from datetime import datetime
from pathlib import Path
from urllib.parse import quote
import json
import sys
import time

import openpyxl

ROOT_DIR = Path(__file__).resolve().parents[2]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.exporter import MERGED_COLUMNS, export_places_to_excel
from src.pc.browser_session import BrowserSession
from src.pc.config import DiagnosticConfig
from src.pc.network_browser_collector import NetworkBrowserCollector, _SEARCH_URL_TEMPLATE
from src.pc.network_list_scraper import dedup_rows as diagnostic_dedup_rows
from src.pc.network_pipeline import run_collection_plan

_COMMON_META = {
    "source_city": "서울특별시",
    "source_district": "강동구",
    "source_layer": "legal_dong",
}
JOBS = [
    {"query": "서울특별시 강동구 천호동 카페", "source_subregion": "천호동", **_COMMON_META},
    {"query": "서울특별시 강동구 길동 카페", "source_subregion": "길동", **_COMMON_META},
    {"query": "서울특별시 강동구 성내동 카페", "source_subregion": "성내동", **_COMMON_META},
    {"query": "서울특별시 강동구 암사동 카페", "source_subregion": "암사동", **_COMMON_META},
]
PER_QUERY_LIMIT = 20
TARGET_COUNT = 50

RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_XLSX_PATH = ROOT_DIR / "output" / f"wire4b_live_50_{RUN_TIMESTAMP}.xlsx"
RESULT_JSON_PATH = Path(__file__).resolve().parent / f"wire4b_live_50_result_{RUN_TIMESTAMP}.json"


def _safe_session_factory():
    # "기본 BrowserSession 경로"를 명시적으로 사용한다(고객용 안전 모드 -
    # visible=False/capture_artifacts=False/keep_open_on_error=False).
    return BrowserSession(DiagnosticConfig.safe_default())


def main() -> int:
    collected_at = datetime.now().strftime("%Y-%m-%d")

    timings: dict = {}
    per_query_log: list = []
    diagnostic_seen: set = set()
    cumulative_unique = 0

    def make_recording_collect_query(collector):
        """collector.collect_query를 그대로 호출하되(동작 변경 없음), 실행 순번/
        진단값을 harness 보고용으로만 기록한다. global unique_added는 진단
        전용 seen set(diagnostic_seen)에만 적용하며 run_collection_plan에는
        전달하지 않는다 - 실제 오케스트레이션 dedup 로직과 완전히 분리된
        관찰용 계산이다."""

        def wrapper(job, per_query_limit):
            nonlocal cumulative_unique
            index = len(per_query_log) + 1
            search_url = _SEARCH_URL_TEMPLATE.format(query=quote(job["query"]))
            t0 = time.perf_counter()
            result = collector.collect_query(job, per_query_limit)
            t1 = time.perf_counter()

            rows = result.get("rows") or []
            newly_unique = diagnostic_dedup_rows(rows, diagnostic_seen)
            cumulative_unique += len(newly_unique)

            per_query_log.append({
                "index": index,
                "query": job["query"],
                "search_url": search_url,
                "wall_seconds": t1 - t0,
                "candidate_response_count": result.get("candidate_response_count"),
                "raw_item_count": result.get("raw_item_count"),
                "local_unique_count": result.get("local_unique_count"),
                "returned_rows_count": len(rows),
                "global_unique_added_diagnostic": len(newly_unique),
                "cumulative_unique_diagnostic": cumulative_unique,
                "parse_error_count": result.get("parse_error_count"),
                "timeout": result.get("timeout"),
                "active_captcha_detected": result.get("active_captcha_detected"),
                "status_429_seen": result.get("status_429_seen"),
                "navigation_error": result.get("navigation_error"),
                "navigation_error_message": (result.get("navigation_error_message") or "")[:200],
            })
            print(f"[wire4b] query {index}/{len(JOBS)} 완료: {job['query']!r} "
                  f"candidate={result.get('candidate_response_count')} raw={result.get('raw_item_count')} "
                  f"local_unique={result.get('local_unique_count')} returned_rows={len(rows)} "
                  f"global_unique_added={len(newly_unique)} cumulative={cumulative_unique} "
                  f"wall={t1 - t0:.2f}s")
            return result

        return wrapper

    print(f"[wire4b] job 수={len(JOBS)}, per_query_limit={PER_QUERY_LIMIT}, target_count={TARGET_COUNT}")
    for i, job in enumerate(JOBS, start=1):
        print(f"[wire4b] job{i}: {job['query']!r}")
    print("[wire4b] live 실행 시작(정확히 1회, 재시도 없음)")

    t_start = time.perf_counter()
    export_result_path = None
    export_error_message = ""
    exported = False

    with NetworkBrowserCollector(collected_at=collected_at, session_factory=_safe_session_factory) as collector:
        t_session_ready = time.perf_counter()
        timings["session_ready_seconds"] = t_session_ready - t_start

        orchestrator_result = run_collection_plan(
            JOBS,
            per_query_limit=PER_QUERY_LIMIT,
            target_count=TARGET_COUNT,
            collected_at=collected_at,
            collect_query=make_recording_collect_query(collector),
        )
        t_orchestrator_done = time.perf_counter()
        timings["orchestrator_seconds"] = t_orchestrator_done - t_session_ready

    t_session_closed = time.perf_counter()
    timings["session_teardown_seconds"] = t_session_closed - t_orchestrator_done

    rows = orchestrator_result.get("rows") or []
    if rows:
        try:
            saved_path = export_places_to_excel(rows, [], [], str(OUTPUT_XLSX_PATH))
            exported = True
            export_result_path = str(saved_path or OUTPUT_XLSX_PATH)
        except Exception as exc:
            export_error_message = f"{type(exc).__name__}: {exc}"
            print(f"[wire4b] Excel 저장 실패: {export_error_message}")
        t_export_done = time.perf_counter()
        timings["export_seconds"] = t_export_done - t_session_closed
    else:
        timings["export_seconds"] = 0.0
        print("[wire4b] rows=0 - export_places_to_excel 호출하지 않음(파일 미생성)")

    excel_verification: dict = {}
    if exported and export_result_path:
        wb = openpyxl.load_workbook(export_result_path)
        merged_ws = wb["통합_결과"]
        headers = [cell.value for cell in merged_ws[1]]
        merged_data_rows = list(merged_ws.iter_rows(min_row=2, values_only=True))
        mobile_ws = wb["원본_모바일"]
        pc_ws = wb["원본_PC"]

        internal_fields = ("place_id", "source_city", "source_district", "source_subregion", "source_layer", "source_query")
        header_leak = [field for field in internal_fields if field in headers]

        def _count_filled(col_name: str):
            if col_name not in headers:
                return None
            idx = headers.index(col_name)
            return sum(1 for row in merged_data_rows if row[idx] not in (None, ""))

        excel_verification = {
            "headers": headers,
            "headers_match_merged_columns": headers == MERGED_COLUMNS,
            "data_row_count": len(merged_data_rows),
            "data_row_count_matches_final_count": len(merged_data_rows) == orchestrator_result.get("final_count"),
            "internal_field_leak": header_leak,
            "mobile_sheet_max_row": mobile_ws.max_row,
            "pc_sheet_max_row": pc_ws.max_row,
            "field_fill_counts": {col: _count_filled(col) for col in MERGED_COLUMNS},
        }

    t_end = time.perf_counter()
    timings["total_wall_seconds"] = t_end - t_start
    executed_count = len(per_query_log)
    timings["avg_query_wall_seconds"] = (
        sum(entry["wall_seconds"] for entry in per_query_log) / executed_count if executed_count else 0.0
    )

    total_returned_rows = sum(entry["returned_rows_count"] for entry in per_query_log)
    total_raw_items = sum((entry["raw_item_count"] or 0) for entry in per_query_log)
    before_trim_count = orchestrator_result.get("before_trim_count") or 0
    global_duplicates_removed = total_returned_rows - before_trim_count
    global_dedup_rate = (global_duplicates_removed / total_returned_rows) if total_returned_rows else 0.0
    diagnostic_matches_orchestrator = cumulative_unique == before_trim_count

    summary = {
        "total_job_count": len(JOBS),
        "per_query_limit": PER_QUERY_LIMIT,
        "target_count": TARGET_COUNT,
        "executed_query_count": orchestrator_result.get("executed_query_count"),
        "skipped_query_count": orchestrator_result.get("skipped_query_count"),
        "before_trim_count": orchestrator_result.get("before_trim_count"),
        "final_count": orchestrator_result.get("final_count"),
        "stop_reason": orchestrator_result.get("stop_reason"),
        "security_blocked": orchestrator_result.get("security_blocked"),
        "status_429_seen": orchestrator_result.get("status_429_seen"),
        "navigation_error": orchestrator_result.get("navigation_error"),
        "navigation_error_message": (orchestrator_result.get("navigation_error_message") or "")[:200],
        "exported": exported,
        "export_path": export_result_path or "",
        "export_error": bool(export_error_message),
        "export_error_message": export_error_message,
        "total_raw_items": total_raw_items,
        "total_returned_rows_after_local_dedup_and_cap": total_returned_rows,
        "global_duplicates_removed": global_duplicates_removed,
        "global_dedup_rate": global_dedup_rate,
        "diagnostic_cumulative_unique_matches_orchestrator_before_trim": diagnostic_matches_orchestrator,
        "timings": timings,
        "per_query_log": per_query_log,
        "excel_verification": excel_verification,
        "retries": 0,
    }

    print("[wire4b] ==== 결과 요약 ====")
    for key, value in summary.items():
        if key in ("timings", "excel_verification", "per_query_log"):
            continue
        print(f"[wire4b] {key}={value}")
    print(f"[wire4b] timings={timings}")
    print(f"[wire4b] excel_verification={excel_verification}")

    RESULT_JSON_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[wire4b] 결과 JSON 저장: {RESULT_JSON_PATH}")

    if exported:
        print(f"[wire4b] Excel 저장 경로: {export_result_path}")

    print("[wire4b] 실행 종료(재실행하지 않음)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
